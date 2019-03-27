# BuildingSync(R), Copyright (c) 2015-2018, Alliance for Sustainable Energy, LLC.
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without modification,
# are permitted provided that the following conditions are met:
#
# (1) Redistributions of source code must retain the above copyright notice, this
#     list of conditions and the following disclaimer.
#
# (2) Redistributions in binary form must reproduce the above copyright notice,
#     this list of conditions and the following disclaimer in the documentation
#     and/or other materials provided with the distribution.
#
# (3) Neither the name of the copyright holder nor the names of any contributors
#     may be used to endorse or promote products derived from this software
#     without specific prior written permission from the respective party.
#
# (4) Other than as required in clauses (1) and (2), distributions in any form of
#     modifications or other derivative works may not use the "BuildingSync"
#     trademark or any other confusingly similar designation without specific
#     prior written permission from Alliance for Sustainable Energy, LLC.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDER(S) AND ANY CONTRIBUTORS "AS
# IS" AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
# IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
# DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER(S), ANY CONTRIBUTORS, THE
# UNITED STATES GOVERNMENT, OR THE UNITED STATES DEPARTMENT OF ENERGY, NOR ANY
# OF THEIR EMPLOYEES, BE LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL,
# EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO,
# PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR
# BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER
# IN CONTRACT, STRICT LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE)
# ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE
# POSSIBILITY OF SUCH DAMAGE.

import openpyxl
import loadxl
import datetime
import os
import warnings
import calendar
import lxml.etree as et
# import xml.etree.ElementTree as et
from xml.dom import minidom

# Known limitations:
# 1) Some of the keys are not scrubbed for those asterisks
# 2) The get*** functions are vulnerable to data loss if keys are not unique
# 3) Order is not preserved in tables, need to use the newer getlistinfo function
# 4) Messing with the labels was probably a mistake
# 5) The Energy Sources table is broken
# 6) There are likely missing refs between the building and measures
# 7) IP/SI is not fully handled
# 8) "All - Delivered Energy" is a bit complicated, only support what's in the template
# 9) Unit conversions, particularly for metered and delivered energy, are too complicated

# This table could be read from the spreadsheet
metered_energy_bsxml_units = {'Electricity': 'kWh',
                              'Natural Gas': 'therms',
                              'Purchased Steam': 'lbs',
                              'Purchased Hot Water': 'MMBtu',
                              'Purchased Chilled Water': 'MMBtu',
                              'Oil': 'Gallons',
                              'Propane': 'Gallons',
                              'Coal': 'Mass ton',
                              'Other': 'kWh',
                              'Thermal - On-Site Generated': 'Unknown',  # This was blank, that doesn't seem right
                              'Electricity - On-Site Generated': 'kWh',
                              'Thermal or Electricity - Exported': 'MMBtu'}
metered_energy_default_units = {'Electricity': 'kWh',
                                'Natural Gas': 'therms',
                                'Purchased Steam': 'lbs District Steam',
                                'Purchased Hot Water': 'MMBtu',
                                'Purchased Chilled Water': 'MMBtu',
                                'Oil': 'gallons (Fuel Oil #2)',
                                'Propane': 'gallons (Propane)',
                                'Coal': 'short ton (coal)',
                                'Other': 'kWh',
                                'Thermal - On-Site Generated': '',  # This is blank, that doesn't seem right
                                'Electricity - On-Site Generated': 'kWh',
                                'Thermal or Electricity - Exported': 'MMBtu'}
# This one is harder to read from the spreadsheet, but could be done
metered_energy_type_lookup = {'Electricity': 'Electricity',
                              'Natural Gas': 'Natural gas',
                              'Purchased Steam': 'District steam',
                              'Purchased Hot Water': 'District hot water',
                              'Purchased Chilled Water': 'District chilled water',
                              'Oil': 'Fuel oil',
                              'Propane': 'Propane',
                              'Coal': 'Coal',
                              'Other': 'Other'}

# This table could be read from the spreadsheet
delivered_energy_default_units = {'Oil': 'Gallons',
                                  'Propane': 'Gallons',
                                  'Coal': 'Mass ton',
                                  'Other': 'kWh'}

bsxml_ballast_lookup = {'Electronic': 'Electronic',
                        'Magnetic': 'Electromagnetic',
                        'N/A': 'No Ballast',
                        'Other': 'Other'}

conversion_to_kBtu = {'kWh': 3.412,
                      'MWh': 3412,
                      'MMBtu': 1000,
                      'therms': 100,
                      'dekatherms': 1000,
                      'MJ': 0.947777778,
                      'cubic feet (NG)': 1.030,
                      'MCF (NG)': 1030000,
                      'lbs District Steam': 1.1940,
                      'lbs District Steam (15 psig) dumped after use': 1.1410,
                      'lbs District Steam (50 psig) dumped after use': 1.1564,
                      'lbs District Steam (140 psig) dumped after use': 1.1721,
                      'lbs (steam, 15 psig) condensate reused': 0.9457,
                      'lbs (steam, 50 psig) condensate reused': 0.9121,
                      'lbs (steam, 140 psig) condensate reused': 0.8619,
                      'short ton (coal)': 19622,
                      'gallons (Fuel Oil #1)': 135,
                      'gallons (Fuel Oil #2)': 139,
                      'gallons (Fuel Oil #3)': 141.8,
                      'gallons (Fuel Oil #4)': 146,
                      'gallons (Fuel Oil #5 Light)': 148,
                      'gallons (Fuel Oil #5 Heavy)': 150,
                      'gallons (Fuel Oil #6)': 154,
                      'gallons (Diesel)': 139,
                      'gallons (Gasoline)': 124,
                      'gallons (Propane)': 92,
                      'cubic feet (Propane)': 2.516}

bsxml_unit_lookup = {'kWh': 'kWh',
                     'MWh': 'MWh',
                     'MMBtu': 'MMBtu',
                     'therms': 'therms',
                     # 'dekatherms': 1000,
                     # 'MJ': 0.947777778,
                     # 'cubic feet (NG)': 1.030,
                     'MCF (NG)': 'MCF',
                     'lbs District Steam': 'lbs',
                     'lbs District Steam (15 psig) dumped after use': 'lbs',
                     'lbs District Steam (50 psig) dumped after use': 'lbs',
                     'lbs District Steam (140 psig) dumped after use': 'lbs',
                     'lbs (steam, 15 psig) condensate reused': 'lbs',
                     'lbs (steam, 50 psig) condensate reused': 'lbs',
                     'lbs (steam, 140 psig) condensate reused': 'lbs',
                     'short ton (coal)': 'Mass ton',
                     'gallons (Fuel Oil #1)': 'Gallons',
                     'gallons (Fuel Oil #2)': 'Gallons',
                     'gallons (Fuel Oil #3)': 'Gallons',
                     'gallons (Fuel Oil #4)': 'Gallons',
                     'gallons (Fuel Oil #5 Light)': 'Gallons',
                     'gallons (Fuel Oil #5 Heavy)': 'Gallons',
                     'gallons (Fuel Oil #6)': 'Gallons',
                     'gallons (Diesel)': 'Gallons',
                     'gallons (Gasoline)': 'Gallons',
                     'gallons (Propane)': 'Gallons'}
# 'cubic feet (Propane)': 2.516}

energysources_labels = ['Energy Source',
                        'ID',
                        None,
                        None,
                        'Type',
                        'Rate schedule']

spacefunctions_211_labels = ['Space Number',
                             'Function type*',
                             'Original intended use',
                             '''Gross Floor Area*
(per space)''',
                             '''Conditioned Area*
(Approx % of total function space)''',
                             'Number of Occupants',
                             'Approximate Plug Loads (W/sf)',
                             'Use (hours/week)',
                             'Use (weeks/year)',
                             'Principal HVAC Type*',
                             'Principal Lighting Type*']

spacefunctions_labels = ['Space Number',
                         'Function type',
                         'Original intended use',
                         'Gross Floor Area',
                         'Percent Conditioned Area',
                         'Number of Occupants',
                         'Approximate Plug Loads (W/sf)',
                         'Use (hours/week)',
                         'Use (weeks/year)',
                         'Principal HVAC Type',
                         'Principal Lighting Type']

L1_eemsummary_header_yi = ['Low-Cost and No-Cost Recommendations',
                           'Modified System',
                           'Impact on Occupant Comfort or IEQ',
                           'Other Non-Energy Impacts',
                           'Cost',
                           'Savings Impact',
                           'Typical ROI',
                           'Priority']

L1_eemsummary_header_er = ['Potential Capital Recommendations',
                           'Modified System',
                           'Impact on Occupant Comfort',
                           'Other Non-Energy Impacts',
                           'Cost',
                           'Savings Impact',
                           'Typical ROI',
                           'Priority']

capacity_units = ["cfh",
                  "ft3/min",
                  "kcf/h",
                  "MCF/day",
                  "gpm",
                  "W",
                  "kW",
                  "hp",
                  "MW",
                  "Btu/hr",
                  "cal/h",
                  "ft-lbf/h",
                  "ft-lbf/min",
                  "Btu/s",
                  "kBtu/hr",
                  "MMBtu/hr",
                  "therms/h",
                  "lbs/h",
                  "Klbs/h",
                  "Mlbs/h",
                  "Cooling ton",
                  "Other"]

L2_equipment_inventory_system_types = ['Heating Plant Type',
                                       'Cooling Plant Type',
                                       'Boiler Type',
                                       'Cooling Delivery Type',
                                       'Heating Delivery Type',
                                       'Heat Recovery Type',
                                       'DX System Type']


def tuple_from_coordinate(coordinate):
    coord = openpyxl.utils.cell.coordinate_from_string(coordinate)
    col = openpyxl.utils.cell.column_index_from_string(coord[0])
    row = coord[1]
    return col, row


class ScanFailure(Exception):
    pass


class LabelMismatch(Exception):
    pass


class MissingRequired(Exception):
    pass


def cellrange(worksheet, mincol=None, minrow=None, maxcol=None, maxrow=None):
    if minrow == maxrow:
        for row in worksheet.iter_rows(min_row=minrow, min_col=mincol,
                                       max_col=maxcol, max_row=maxrow):
            return [el.value for el in row]
    elif mincol == maxcol:
        for col in worksheet.iter_cols(min_row=minrow, min_col=mincol,
                                       max_col=maxcol, max_row=maxrow):
            return [el.value for el in col]
    results = []
    for row in worksheet.iter_rows(min_row=minrow, min_col=mincol,
                                   max_col=maxcol, max_row=maxrow):
        results.append([el.value for el in row])
    return results


def getlabeledvalues(worksheet, cellrange, labelcolor=0, IP=True,
                     valuecolor=8, variablelength=False, hasunits=False):
    if isinstance(cellrange, str):
        try:
            rangetuple = openpyxl.utils.range_boundaries(cellrange)
        except TypeError:
            raise TypeError('Unable to determine cell range')
    else:
        if len(cellrange) == 4:
            rangetuple = cellrange
        else:
            raise TypeError('Unable to determine cell range')
    labelcol = min(rangetuple[0], rangetuple[2])
    valuecol = max(rangetuple[0], rangetuple[2])
    if rangetuple[1] < rangetuple[3]:
        minrow = rangetuple[1]
        maxrow = rangetuple[3]
    else:
        maxrow = rangetuple[1]
        minrow = rangetuple[3]
    result = {}
    if not hasunits:
        for row in worksheet.iter_rows(min_row=minrow, min_col=labelcol,
                                       max_col=valuecol, max_row=maxrow):
            if row[0].value != None and row[-1].value != None:
                if variablelength:
                    if (row[0].fill.start_color.index != labelcolor or
                            row[-1].fill.start_color.index != valuecolor):
                        break
                result[row[0].value] = row[-1].value
    else:
        for row in worksheet.iter_rows(min_row=minrow, min_col=labelcol,
                                       max_col=valuecol + 1, max_row=maxrow):
            if row[0].value != None and row[-2].value != None:
                if variablelength:
                    if (row[0].fill.start_color.index != labelcolor or
                            row[-2].fill.start_color.index != valuecolor):
                        break
                if row[-1].value != None:
                    # Handle the units, this could get ugly
                    units = row[-1].value

                    if units == '=IF(Instructions!$B$18="IP","sq ft","sq m")':
                        if IP:
                            units = 'sq ft'
                        else:
                            units = 'sq m'
                    key = row[0].value.rstrip() + (' (%s)' % units)
                    result[key] = row[-2].value
                else:
                    result[row[0].value] = row[-2].value
    return result


def getlist(worksheet, cellrange, variablelength=False, fillcolor=8):
    if isinstance(cellrange, str):
        try:
            rangetuple = openpyxl.utils.range_boundaries(cellrange)
        except TypeError:
            raise TypeError('Unable to determine cell range')
    else:
        if len(cellrange) == 4:
            rangetuple = cellrange
        else:
            raise TypeError('Unable to determine cell range')
    diff = (rangetuple[2] - rangetuple[0],
            rangetuple[3] - rangetuple[1])
    result = []
    if diff[0] == 0:
        listcol = rangetuple[0]
        for row in worksheet.iter_rows(min_row=rangetuple[1], min_col=listcol,
                                       max_col=listcol, max_row=rangetuple[3]):
            if variablelength:
                if not row[0].value or row[0].fill.start_color.index != fillcolor:
                    break
            result.append(row[0].value)
    elif diff[1] == 0:
        listrow = rangetuple[1]
        for col in worksheet.iter_cols(min_col=rangetuple[0], min_row=listrow,
                                       max_row=listrow, max_col=rangetuple[2]):
            if variablelength:
                if not col[0].value or col[0].fill.start_color.index != fillcolor:
                    break
            result.append(col[0])
    return result


def getinfo(worksheet, cellrange, variablelength=False, fillcolor=8,
            labels=None, inrows=True, keepemptyrows=False, keepemptycells=True):
    if isinstance(cellrange, str):
        try:
            rangetuple = openpyxl.utils.range_boundaries(cellrange)
        except TypeError:
            raise TypeError('Unable to determine cell range')
    else:
        if len(cellrange) == 4:
            rangetuple = cellrange
        else:
            raise TypeError('Unable to determine cell range')
    result = {}
    if inrows:
        listcol = rangetuple[0]
        for row in worksheet.iter_rows(min_col=rangetuple[0], min_row=rangetuple[1],
                                       max_col=rangetuple[2], max_row=rangetuple[3]):
            if variablelength:
                if (not row[0].value
                        or row[0].fill.start_color.index != fillcolor):
                    break
            elif not keepemptyrows:
                if not row[0].value:
                    continue
            data = [el.value for el in row[1:]]
            if not keepemptyrows:
                count = 0
                for el in data:
                    if el:
                        count += 1
                if count == 0:
                    continue
            if labels:
                if keepemptycells:
                    data = dict(zip(labels[1:], data))
                else:
                    data = dict([el for el in zip(labels[1:], data) if el[1] != None])
            result[row[0].value] = data
    else:
        listrow = rangetuple[1]
        for col in worksheet.iter_cols(min_col=rangetuple[0], min_row=rangetuple[1],
                                       max_row=rangetuple[3], max_col=rangetuple[2]):
            if variablelength:
                if (not col[0].value
                        or col[0].fill.start_color.index != fillcolor):
                    break
            elif not keepemptyrows:
                if not col[0].value:
                    continue
            data = [el.value for el in col[1:]]
            if not keepemptyrows:
                count = 0
                for el in data:
                    if el:
                        count += 1
                if count == 0:
                    continue
            if labels:
                if keepemptycells:
                    data = dict(zip(labels[1:], data))
                else:
                    data = dict([el for el in zip(labels[1:], data) if el[1] != None])
            result[col[0].value] = data
    return result


def gettable(worksheet, cellrange, variablelength=False, fillcolor=8,
             labels=None, inrows=True, keepempty=False):
    if isinstance(cellrange, str):
        try:
            rangetuple = openpyxl.utils.range_boundaries(cellrange)
        except TypeError:
            raise TypeError('Unable to determine cell range')
    else:
        if len(cellrange) == 4:
            rangetuple = cellrange
        else:
            raise TypeError('Unable to determine cell range')
    result = []
    if inrows:
        listcol = rangetuple[0]
        for row in worksheet.iter_rows(min_col=rangetuple[0], min_row=rangetuple[1],
                                       max_col=rangetuple[2], max_row=rangetuple[3]):
            if variablelength:
                if (not row[0].value
                        or row[0].fill.start_color.index != fillcolor):
                    break
            elif not keepempty:
                if not row[0].value:
                    continue
            data = [el.value for el in row]
            if not keepempty:
                count = 0
                for el in data:
                    if el:
                        count += 1
                if count == 0:
                    continue
            if labels:
                data = dict(zip(labels, data))
            result.append(data)
    else:
        listrow = rangetuple[1]
        for col in worksheet.iter_cols(min_col=rangetuple[0], min_row=rangetuple[1],
                                       max_row=rangetuple[3], max_col=rangetuple[2]):
            if variablelength:
                if (not col[0].value
                        or col[0].fill.start_color.index != fillcolor):
                    break
            elif not keepempty:
                if not col[0].value:
                    continue
            data = [el.value for el in col]
            if not keepempty:
                count = 0
                for el in data:
                    if el:
                        count += 1
                if count == 0:
                    continue
            if labels:
                data = dict(zip(labels, data))
            result.append(data)
    return result


def getlistinfo(worksheet, cellrange, variablelength=False, fillcolor=8,
                labels=None, inrows=True, keepempty=False):
    if isinstance(cellrange, str):
        try:
            rangetuple = openpyxl.utils.range_boundaries(cellrange)
        except TypeError:
            raise TypeError('Unable to determine cell range')
    else:
        if len(cellrange) == 4:
            rangetuple = cellrange
        else:
            raise TypeError('Unable to determine cell range')
    result = []
    if inrows:
        listcol = rangetuple[0]
        for row in worksheet.iter_rows(min_col=rangetuple[0], min_row=rangetuple[1],
                                       max_col=rangetuple[2], max_row=rangetuple[3]):
            if variablelength:
                if (not row[0].value
                        or row[0].fill.start_color.index != fillcolor):
                    break
            elif not keepempty:
                if not row[0].value:
                    continue
            data = [el.value for el in row]

            if not keepempty:
                count = 0
                for el in data:
                    if el:
                        count += 1
                if count == 0:
                    continue
            if labels:
                # Have to handle None in the labels
                out = {}
                for i in range(len(labels)):
                    if labels[i] == None or data[i] == None:
                        continue
                    out[labels[i]] = data[i]
                data = out
                # data = dict(zip(labels,data))
            result.append(data)
    else:
        listrow = rangetuple[1]
        for col in worksheet.iter_cols(min_col=rangetuple[0], min_row=rangetuple[1],
                                       max_row=rangetuple[3], max_col=rangetuple[2]):
            if variablelength:
                if (not col[0].value
                        or col[0].fill.start_color.index != fillcolor):
                    break
            elif not keepempty:
                if not col[0].value:
                    continue
            data = [el.value for el in col[1:]]
            if not keepempty:
                count = 0
                for el in data:
                    if el:
                        count += 1
                if count == 0:
                    continue
            if labels:
                # Have to handle None in the labels
                out = {}
                for i in range(len(labels)):
                    if labels[i] == None or data[i] == None:
                        continue
                    out[labels[i]] = data[i]
                data = out
                # data = dict(zip(labels,data))
            result.append(data)
    return result


def gettabular(worksheet, mincol, minrow, maxcol, maxrow):
    results = []
    for row in worksheet.iter_rows(min_row=minrow, min_col=mincol,
                                   max_col=maxcol, max_row=maxrow):
        results.append([el.value for el in row])
    return results


def getcellrange(worksheet, cellrange):
    try:
        mincol, minrow, maxcol, maxrow = openpyxl.utils.range_boundaries(cellrange)
    except TypeError:
        return None
    return gettabular(worksheet, mincol, minrow, maxcol, maxrow)


def scanRowForEmpty(worksheet, col, row, maxcol=256):
    count = 0
    for col in worksheet.iter_cols(min_col=col, min_row=row,
                                   max_row=row):
        if not col[0].value:
            return count
        count += 1


def scanForExpandableColumnTable(worksheet, mincol=1, minrow=1, maxrow=1,
                                 minentries=1):
    result = []
    for col in worksheet.iter_cols(min_col=mincol, min_row=minrow,
                                   max_row=maxrow):
        count = 0
        data = []
        for el in col:
            data.append(el.value)
            if el.value:
                count += 1
        if count < minentries:
            return result
        result.append(data)
    return result


def scanForHeaderRow(worksheet, mincol, minrow, header):
    maxcol = mincol + len(header) - 1
    count = 0
    for row in worksheet.iter_rows(min_col=mincol, min_row=minrow,
                                   max_col=maxcol):
        data = [el.value for el in row]
        if data == header:
            return minrow + count
        count += 1
    raise ScanFailure('Failed to find header')


def scan_for_cell_value(worksheet, mincol=None, minrow=None, maxcol=None,
                        maxrow=None, value=None):
    for row in worksheet.iter_rows(min_col=mincol, min_row=minrow,
                                   max_col=maxcol, max_row=maxrow):
        for el in row:
            if el.value == value:
                return tuple_from_coordinate(el.coordinate)
    raise ScanFailure('Failed to find cell value')


def read_all_building(worksheet):
    '''Read the 'All - Building' sheet
    
    The first several items are fixed in size and location, but
    the "Space Function" table looks to be expandable. Everything
    after that needs to be found.
    '''
    # High level building information
    bldg_info = getlabeledvalues(worksheet, 'A3:B13')
    bldg_info.update(getlabeledvalues(worksheet, 'A19:B25'))
    bldg_info.update(getlabeledvalues(worksheet, 'E15:F22'))
    # Scrub any dates
    for key, value in bldg_info.items():
        if isinstance(value, datetime.datetime):
            bldg_info[key] = str(value)
    # Excluded space
    excluded_spaces = getlist(worksheet, 'E24:E26', variablelength=True)
    # Space Function
    space_function = getlabeledvalues(worksheet, 'A29:B33', variablelength=True,
                                      labelcolor=8)
    # Occupancy
    # Look for Occupancy*
    cellcol, cellrow = scan_for_cell_value(worksheet, mincol=1, minrow=34,
                                           maxcol=1, value='Occupancy*')
    occupancy = getlabeledvalues(worksheet, [cellcol, cellrow + 1,
                                             cellcol, cellrow + 5],
                                 hasunits=True)
    # Energy Sources
    # Look for Energy Sources**
    cellcol, cellrow = scan_for_cell_value(worksheet, mincol=1, minrow=41,
                                           maxcol=1, value='Energy Sources**')
    cellrow += 1
    # Check the labels
    labels = cellrange(worksheet, mincol=cellcol, minrow=cellrow,
                       maxcol=cellcol + 5, maxrow=cellrow)
    if labels != energysources_labels:
        raise LabelMismatch('Mismatch in energy sources labels')
    cellrow += 1
    energy_sources = getlistinfo(worksheet, [cellcol, cellrow, cellcol + 5, None],
                                 variablelength=True, labels=labels,
                                 inrows=True, keepempty=False)

    # Facility Description
    # Look for Facility Description - Notable Conditions
    cellcol, cellrow = scan_for_cell_value(worksheet, mincol=1, minrow=54, maxcol=1,
                                           value='Facility Description - Notable Conditions')
    description = worksheet.cell(column=cellcol, row=cellrow + 1).value

    # Package the data
    bldg_info['Occupancy'] = occupancy
    bldg_info['Energy Sources'] = energy_sources
    bldg_info['Facility Description'] = description
    bldg_info['Space Function'] = space_function
    bldg_info['Excluded Spaces'] = excluded_spaces

    return bldg_info


def read_utility_table(worksheet, name, labels, row=1, col=1):
    # Scan for the name
    cellcol, cellrow = scan_for_cell_value(worksheet, mincol=col, minrow=row,
                                           maxcol=col,
                                           value=name)
    # Header is 3 rows down, data 1 more down
    cellrow += 4

    data = gettable(worksheet, [cellcol, cellrow,
                                cellcol + len(labels), None],
                    variablelength=True, inrows=True, labels=labels)

    return data


def read_utility_definition(worksheet, name, row=1, col=1):
    # Scan for the name
    cellcol, cellrow = scan_for_cell_value(worksheet, mincol=col, minrow=row,
                                           maxcol=col,
                                           value=name + ': Definition')
    # Labeled items are one row down
    cellrow += 1
    data = getlabeledvalues(worksheet, [cellcol, cellrow,
                                        cellcol + 1, cellrow + 1])

    return data


def read_all_metered_energy(worksheet):
    '''Read the 'All - Metered Energy' sheet

    This sheet has at least three utilities, each with two tables
    '''
    # There's a lot of hardcoding here, it is not clear how this sheet
    # would be expanded
    header_info = getlabeledvalues(worksheet, 'A5:C8')
    methodology = worksheet['A12'].value
    electricity_labels = ['Start Date', 'End Date', 'Days', 'Use', 'Peak', 'Cost']
    other_labels = ['Start Date', 'End Date', 'Days', 'Use', 'Cost']
    data = {}
    if 'Utility #1' in header_info:
        # This one is supposed to be electricity if it is present. It is not
        # clear if it can be something else if electricity is present.
        utility = read_utility_table(worksheet, 'Utility #1',
                                     electricity_labels,
                                     row=15, col=1)
        definition = read_utility_definition(worksheet, 'Utility #1',
                                             row=15, col=1)
        data['Utility #1'] = {}
        data['Utility #1']['Data'] = utility
        data['Utility #1']['Definition'] = definition
        data['Utility #1']['Type'] = 'Electricity'  # Here's where we could check if this is true
    if 'Utility #2' in header_info:
        utility = read_utility_table(worksheet, 'Utility #2',
                                     other_labels,
                                     row=15, col=1)
        definition = read_utility_definition(worksheet, 'Utility #2',
                                             row=15, col=1)
        data['Utility #2'] = {}
        data['Utility #2']['Data'] = utility
        data['Utility #2']['Definition'] = definition
        data['Utility #2']['Type'] = header_info['Utility #2']
    if 'Utility #3' in header_info:
        utility = read_utility_table(worksheet, 'Utility #3',
                                     other_labels,
                                     row=15, col=1)
        definition = read_utility_definition(worksheet, 'Utility #3',
                                             row=15, col=1)
        data['Utility #3'] = {}
        data['Utility #3']['Data'] = utility
        data['Utility #3']['Definition'] = definition
        data['Utility #3']['Type'] = header_info['Utility #3']
    return data


def read_all_delivered_energy(worksheet):
    '''Read the 'All - Delivered Energy' sheet

    This sheet has at least one set of three tables, right now only support one set
    '''
    header_info = getlabeledvalues(worksheet, 'C2:D4')

    # Look for "Delivery date" in the second column
    cellcol, cellrow = scan_for_cell_value(worksheet, mincol=2, minrow=5,
                                           maxcol=2,
                                           value='Delivery date')
    # Get the table
    labels = ['Delivery date', 'Volume', 'kBTU', 'Cost']
    delivered = gettable(worksheet, [cellcol, cellrow + 1,
                                     cellcol + 3, None],
                         variablelength=True,
                         inrows=True, labels=labels)

    # Lood for 'Estimated Annual Use**' in the first column
    cellcol, cellrow = scan_for_cell_value(worksheet, mincol=1, minrow=cellrow + 1,
                                           maxcol=1,
                                           value='Estimated Annual Use**')
    estimated_annual_use = worksheet.cell(column=cellcol + 2, row=cellrow).value
    header_info['Estimated Annual Use**'] = estimated_annual_use
    return {'Definition': header_info, 'Data': delivered}


def read_L1_eem_summary(worksheet):
    '''Read the 'L1 - EEM Summary' sheet

    This sheet is apparently two tables. Find one and then the other.
    '''
    # Find the first table
    row0 = scanForHeaderRow(worksheet, 1, 3, L1_eemsummary_header_yi)

    # Find the second table
    row1 = scanForHeaderRow(worksheet, 1, row0 + 1, L1_eemsummary_header_er)

    # Get the first table
    lowcost = getinfo(worksheet, [1, row0 + 1, len(L1_eemsummary_header_yi), row1 - 1],
                      labels=L1_eemsummary_header_yi)
    # Get the second table
    potentialcapital = getinfo(worksheet, [1, row1 + 1, len(L1_eemsummary_header_er), None],
                               labels=L1_eemsummary_header_er)

    return {'Low-Cost and No-Cost Recommendations': lowcost,
            'Potential Capital Recommendations': potentialcapital}


def read_space_functions(worksheet):
    '''Read the 'All - Space Functions' sheet

    This sheet is basically one big table.
    '''
    # Look for "Space Number" in the first column
    cellcol, cellrow = scan_for_cell_value(worksheet, mincol=1, minrow=1, maxcol=1,
                                           value='Space Number')

    # Check the labels
    labels = cellrange(worksheet, mincol=cellcol, minrow=cellrow, maxcol=cellcol,
                       maxrow=cellrow + len(spacefunctions_211_labels) - 1)
    if labels != spacefunctions_211_labels:
        raise LabelMismatch('Mismatch in space function labels')
    space_functions = getinfo(worksheet, [cellcol + 1, cellrow, None,
                                          cellrow + len(spacefunctions_211_labels) - 1],
                              inrows=False, labels=spacefunctions_labels)
    return space_functions


def handle_key_formulas(key, IP):
    newkey = key
    if '&IF(Instructions!B18="IP","(ft2)","m2")' in key:
        unit = {True: ' (ft2)', False: ' (m2)'}[IP]
        newkey = newkey[1:].replace('&IF(Instructions!B18="IP","(ft2)","m2")', '')
        newkey = newkey.replace('"', '')
        newkey = newkey.strip() + unit
    return newkey


def read_L2_envelope(worksheet, IP=True):
    '''Read the 'L2 - Envelope' sheet

    This sheet is a combination of free entry, one choice, and checkboxes
    '''
    # Get the top info
    info = getlabeledvalues(worksheet, 'A3:B6', hasunits=True, IP=IP)
    info.update(getlabeledvalues(worksheet, 'A7:B10'))
    info.update(getlabeledvalues(worksheet, 'E12:F13'))
    info.update(getlabeledvalues(worksheet, 'A15:B15'))
    rvalues = getcellrange(worksheet, 'E3:F5')
    info['Total exposed above grade wall area R value'] = rvalues[0][1]
    info['Below grade wall area R value'] = rvalues[1][1]
    info['Roof area R value'] = rvalues[2][1]
    keys = list(info.keys())
    for key in keys:
        if key.startswith('='):
            newkey = handle_key_formulas(key, IP)
            value = info.pop(key, None)
            info[newkey] = value
    # Ye olde awful tables
    roof_constuction_table = ['Check Box 1',
                              'Check Box 2',
                              'Check Box 3',
                              'Check Box 4',
                              'Check Box 5',
                              'Check Box 6']
    fenestration_frame_types_table = ['Check Box 7',
                                      'Check Box 8',
                                      'Check Box 9',
                                      'Check Box 40',
                                      'Check Box 10']
    floor_construction_table = ['Check Box 11',
                                'Check Box 12',
                                'Check Box 13',
                                'Check Box 14',
                                'Check Box 15']
    fenestration_glass_types_table = ['Check Box 16',
                                      'Check Box 17',
                                      'Check Box 18',
                                      'Check Box 19',
                                      'Check Box 26',
                                      'Check Box 37']
    wall_constructions_table = ['Check Box 20',
                                'Check Box 21',
                                'Check Box 22',
                                'Check Box 23',
                                'Check Box 24',
                                'Check Box 27',
                                'Check Box 36']
    foundation_type_table = ['Check Box 28',
                             'Check Box 29',
                             'Check Box 30',
                             'Check Box 31',
                             'Check Box 32']
    big_table = {'Roof Construction': roof_constuction_table,
                 'Fenestration Frame Types': fenestration_frame_types_table,
                 'Floor Construction': floor_construction_table,
                 'Fenestration Glass Types': fenestration_glass_types_table,
                 'Wall Constructions': wall_constructions_table,
                 'Foundation Type': foundation_type_table}
    # Try to use that stuff now
    for tablename, checkboxes in big_table.items():
        table = []
        for name in checkboxes:
            if worksheet.controls[name].checked:
                table.append(worksheet.controls[name].text)
        if table:
            info[tablename] = table
    return info


def read_L2_hvac(worksheet):
    zone_controls_table = ['Check Box 73',
                           'Check Box 69',
                           'Check Box 67',
                           'Check Box 68']
    central_plant_controls_table = ['Check Box 72',
                                    'Check Box 77',
                                    'Check Box 76',
                                    'Check Box 103']
    heat_recovery_table = ['Check Box 81',
                           'Check Box 82']
    outside_air_table = ['Check Box 79',
                         'Check Box 78',
                         'Check Box 84',
                         'Check Box 86']
    exhaust_fans_none = 'Check Box 91'
    exhaust_fans_table = ['Check Box 92',
                          'Check Box 102']
    cooling_dist_equip_table = ['Check Box 1',
                                'Check Box 2',
                                'Check Box 3',
                                'Check Box 4',
                                'Check Box 5',
                                'Check Box 87',
                                'Check Box 88',
                                'Check Box 7',
                                'Check Box 6']
    heating_dist_equip_table = ['Check Box 59',
                                'Check Box 60',
                                'Check Box 61',
                                'Check Box 62',
                                'Check Box 63',
                                'Check Box 64',
                                'Check Box 65']
    cooling_source_none = 'Check Box 8'
    cooling_source_table = ['Check Box 9',
                            'Check Box 10',
                            'Check Box 11',
                            'Check Box 85',
                            'Check Box 13']
    cooling_source_other = 'TextBox 89'
    chiller_input_table = ['Check Box 25',
                           'Check Box 26',
                           'Check Box 27',
                           'Check Box 28',
                           'Check Box 53',
                           'Check Box 52',
                           'Check Box 54']
    compressor_table = ['Check Box 29',
                        'Check Box 31',
                        'Check Box 33',
                        'Check Box 55']
    condenser_table = ['Check Box 30',
                       'Check Box 32',
                       'Check Box 58',
                       'Check Box 56',
                       'Check Box 57']
    heating_source_none = 'Check Box 14'
    heating_source_table = ['Check Box 15',
                            'Check Box 16',
                            'Check Box 17',
                            'Check Box 18',
                            'Check Box 89']
    heating_source_other = 'TextBox 91'
    heating_fuel_table = ['Check Box 34',
                          'Check Box 35',
                          'Check Box 36',
                          'Check Box 37']
    heating_fuel_oil_grade = 'TextBox 88'
    boiler_type_table = ['Check Box 42',
                         'Check Box 43',
                         'Check Box 38',
                         'Check Box 39']
    shw_dhw_source_table = ['Check Box 20',
                            'Check Box 44',
                            'Check Box 21',  #
                            'Check Box 22',  #
                            'Check Box 47',
                            'Check Box 45',  #
                            'Check Box 46',  #
                            'Check Box 23',
                            'Check Box 24']
    shw_dhw_fuel_table = ['Check Box 48',
                          'Check Box 49',
                          'Check Box 50',
                          'Check Box 51']
    shw_dhw_fuel_oil_grade = 'TextBox 1'
    shw_dhw_fuel_other = 'TextBox 87'

    info = {}
    big_table = {'Zone Controls': zone_controls_table,
                 'Central Plant Controls': central_plant_controls_table,
                 'Outside Air': outside_air_table,
                 'Heat Recovery': heat_recovery_table,
                 'Cooling Distribution Equipment Type': cooling_dist_equip_table,
                 'Heating Distribution Equipment Type': heating_dist_equip_table,
                 'Chiller Input': chiller_input_table,
                 'Compressor': compressor_table,
                 'Condenser': condenser_table,
                 'Heating Fuel': heating_fuel_table,
                 'Boiler Type': boiler_type_table,
                 'SHW/DHW Source': shw_dhw_source_table,
                 'SHW/DHW Fuel': shw_dhw_fuel_table}
    check_table = {'Exhaust Fans': (exhaust_fans_none, exhaust_fans_table),
                   'Cooling Source': (cooling_source_none, cooling_source_table),
                   'Heating Source': (heating_source_none, heating_source_table)}
    for tablename, checkboxes in big_table.items():
        table = []
        for name in checkboxes:
            prefix = ''
            if name in ['Check Box 21', 'Check Box 22']:
                prefix = 'Indirect fired - '
            elif name in ['Check Box 45', 'Check Box 46']:
                prefix = 'Direct fired - '
            if worksheet.controls[name].checked:
                table.append(prefix + worksheet.controls[name].text)
        if table:
            info[tablename] = table
    for tablename, value in check_table.items():
        table = []
        keep, checkboxes = value
        if worksheet.controls[keep].checked:
            for name in checkboxes:
                if worksheet.controls[name].checked:
                    table.append(worksheet.controls[name].text)
        if table:
            info[tablename] = table
    # Handle the entry textboxes
    if 'Cooling Source' in info:
        for i, el in enumerate(info['Cooling Source']):
            if el.startswith('Other'):
                if cooling_source_other in worksheet.textboxes:
                    info['Cooling Source'][i] = 'Other (%s)' % worksheet.textboxes[cooling_source_other]
                else:
                    info['Cooling Source'][i] = 'Other (Unspecified)'
    if 'Heating Fuel' in info:
        for i, el in enumerate(info['Heating Fuel']):
            if el.startswith('Oil'):
                if heating_fuel_oil_grade in worksheet.textboxes:
                    info['Heating Fuel'][i] = 'Oil (%s)' % worksheet.textboxes[heating_fuel_oil_grade]
                else:
                    info['Heating Fuel'][i] = 'Oil (Unspecified Grade)'
    if 'SHW/DHW Fuel' in info:
        for i, el in enumerate(info['SHW/DHW Fuel']):
            if el.startswith('Oil'):
                if shw_dhw_fuel_oil_grade in worksheet.textboxes:
                    info['SHW/DHW Fuel'][i] = 'Oil (%s)' % worksheet.textboxes[shw_dhw_fuel_oil_grade]
                else:
                    info['SHW/DHW Fuel'][i] = 'Oil (Unspecified Grade)'
            if el.startswith('Other'):
                if shw_dhw_fuel_other in worksheet.textboxes:
                    info['SHW/DHW Fuel'][i] = 'Other (%s)' % worksheet.textboxes[shw_dhw_fuel_other]
                else:
                    info['SHW/DHW Fuel'][i] = 'Other (Unspecified)'
    return info


def read_L2_equipment_inventory(worksheet):
    # Look for "ID" in the first column
    cellcol, cellrow = scan_for_cell_value(worksheet, mincol=1, minrow=2,
                                           maxcol=1,
                                           value='ID')
    # The labels we expect
    expected = ['ID', 'Description', 'Location', 'Type', 'Units',
                'Rated efficiency (as applicable)', 'Output Capacity',
                'Area Served', 'Approx Year Installed', 'Condition       (excellent, good, average, poor)']

    # Check the labels
    labels = cellrange(worksheet, mincol=cellcol, minrow=cellrow,
                       maxcol=cellcol + len(expected) - 1,
                       maxrow=cellrow)
    if labels != expected:
        raise LabelMismatch('Mismatch in equipment inventory labels')
    inventory = getinfo(worksheet, [cellcol, cellrow + 1,
                                    cellcol + len(labels) - 1,
                                    None], variablelength=True,
                        inrows=True, keepemptycells=False, labels=labels)
    return inventory


lighting_sources_labels = ['Lighting Source Type(s)',
                           'Ballast Type(s)',
                           'Control(s)',
                           'Space Type(s)*',
                           'Approx % Area Served']

load_labels = ['Major Process/Plug Load Type(s)**',
               'Key Operational Details***']


def read_L2_lighting(worksheet):
    '''Read the 'L2 - Lighting Elec & Plug Loads' sheet

    This sheet is two tables.
    '''
    # Look for "Lighting Source Type(s)" in the first column
    cellcol, cellrow = scan_for_cell_value(worksheet, mincol=1, minrow=1, maxcol=1,
                                           value='Lighting Source Type(s)')

    # Check the labels
    labels = cellrange(worksheet, mincol=cellcol, minrow=cellrow,
                       maxcol=cellcol + len(lighting_sources_labels) - 1,
                       maxrow=cellrow)
    if labels != lighting_sources_labels:
        raise LabelMismatch('Mismatch in lighting source labels')
    lighting = getinfo(worksheet, [cellcol, cellrow + 1,
                                   cellcol + len(lighting_sources_labels) - 1,
                                   None], variablelength=True,
                       inrows=True, labels=lighting_sources_labels)
    # Lood for 'Major Process/Plug Load Type(s)**' in the first column
    # This is another table with merged columns
    cellcol, cellrow = scan_for_cell_value(worksheet, mincol=1, minrow=cellrow + 1,
                                           maxcol=1,
                                           value='Major Process/Plug Load Type(s)**')
    labels = cellrange(worksheet, mincol=cellcol, minrow=cellrow,
                       maxcol=cellcol + len(load_labels) - 1,
                       maxrow=cellrow)
    if labels != load_labels:
        raise LabelMismatch('Mismatch in process/plug load labels')
    loads = getinfo(worksheet, [cellcol, cellrow + 1,
                                cellcol + len(load_labels) - 1,
                                None], variablelength=True,
                    inrows=True, labels=load_labels)
    return {'Lighting Source Type(s)': lighting,
            'Major Process/Plug Load Type(s)**': loads}


def read_L2_eem_summary(worksheet):
    labels = ['Description', 'Energy Cost Savings', 'Non-energy Cost Savings', 'Peak Demand Savings (kW)',
              'Utility #1', 'Utility #2', 'Utility #3', 'Delivered Energy',
              'Measure Cost', 'Potential Incentives', 'Measure Life (years)']
    # Find 'Low-Cost and No-Cost Recommendations'
    col0, row0 = scan_for_cell_value(worksheet, mincol=1, minrow=5,
                                     maxcol=1, value='Low-Cost and No-Cost Recommendations')
    # Find 'Potential Capital Recommendations'
    col1, row1 = scan_for_cell_value(worksheet, mincol=1, minrow=14,
                                     maxcol=1, value='Potential Capital Recommendations')
    # Find 'Potential Capital Recommendations'
    col2, row2 = scan_for_cell_value(worksheet, mincol=1, minrow=15,
                                     maxcol=1, value='TOTALS (Recommended Measures)')
    # Get the first table
    lowcost = getinfo(worksheet, [0, row0 + 1, len(labels), row1 - 1], labels=labels)
    # Get the second table
    potentialcapital = getinfo(worksheet, [0, row1 + 1, len(labels), row2 - 1],
                               labels=labels)

    return {'Low-Cost and No-Cost Recommendations': lowcost,
            'Potential Capital Recommendations': potentialcapital}


def read_std211_xlsx(workbook, IP=True):
    ''' Read Standard 211 information from an Excel workbook into a dictionary '''
    std211 = {}
    #
    # Read the 'All - Building' sheet
    std211['All - Building'] = read_all_building(workbook['All - Building'])
    # Read the 'All - Metered Energy' sheet
    std211['All - Metered Energy'] = read_all_metered_energy(workbook['All - Metered Energy'])
    # Read the 'All - Delivered Energy' sheet
    std211['All - Delivered Energy'] = read_all_delivered_energy(workbook['All - Delivered Energy'])
    # Read the 'All - Space Functions' sheet
    std211['All - Space Functions'] = read_space_functions(workbook['All - Space Functions'])
    # Read the 'L1 - EEM Summary' sheet
    std211['L1 - EEM Summary'] = read_L1_eem_summary(workbook['L1 - EEM Summary'])
    # Read the 'L2 - Envelope' sheet
    std211['L2 - Envelope'] = read_L2_envelope(workbook['L2 - Envelope'], IP=IP)
    # Read the 'L2 - HVAC' sheet
    std211['L2 - HVAC'] = read_L2_hvac(workbook['L2 - HVAC'])
    # Read the 'L2 - Equipment Inventory' sheet
    std211['L2 - Equipment Inventory'] = read_L2_equipment_inventory(workbook['L2 Equipment Inventory'])
    # Read the 'L2 - Lighting Elec & Plug Loads' sheet
    std211['L2 - Lighting Elec & Plug Loads'] = read_L2_lighting(workbook['L2 - Lighting Elec & Plug Loads'])
    # Read the 'L2 - EEM Summary' sheet
    std211['L2 - EEM Summary'] = read_L2_eem_summary(workbook['L2 - EEM Summary'])
    # Done!
    return std211


def process_zip(pc):
    separators = ['-', ' ']
    for sep in separators:
        five, guido, four = pc.partition('-')
        five = five.strip()
        four = four.strip()
        if len(five) != 5 or len(four) != 4:
            continue
        # Check for numbers?
        return five, four
    return pc, None


def determine_frequency(start, end):
    frequency = 'Other'
    delta = end - start
    leapadd = 0
    quarteradd = 0
    try:
        frequency = {60: '1 minute',
                     600: '10 minute',
                     900: '15 minute',
                     1800: '30 minute',
                     3600: 'Hour',
                     86400: 'Day',
                     604800: 'Week'}[delta.total_seconds()]
    except KeyError:
        if start.month == 1:
            if calendar.isleap(start.year):
                leapadd = 1
            try:
                frequency = {30: 'Month',
                             31: 'Month',
                             89 + leapadd: 'Quarter',
                             90 + leapadd: 'Quarter',
                             364 + leapadd: 'Annual',
                             365 + leapadd: 'Annual'}[delta.days]
            except KeyError:
                pass
        if start.month == 2:
            if calendar.isleap(start.year):
                if start.day < 29:
                    leapadd = 1
            try:
                frequency = {27 + leapadd: 'Month',
                             28 + leapadd: 'Month',
                             88 + leapadd: 'Quarter',
                             89 + leapadd: 'Quarter',
                             364 + leapadd: 'Annual',
                             365 + leapadd: 'Annual'}[delta.days]
            except KeyError:
                pass
        elif start.month in [3, 5, 7, 8]:
            if calendar.isleap(start.year + 1):
                leapadd = 1
            try:
                frequency = {30: 'Month',
                             31: 'Month',
                             91: 'Quarter',
                             92: 'Quarter',
                             364 + leapadd: 'Annual',
                             365 + leapadd: 'Annual'}[delta.days]
            except KeyError:
                pass
        elif start.month in [4, 9]:
            if calendar.isleap(start.year + 1):
                leapadd = 1
            try:
                frequency = {29: 'Month',
                             30: 'Month',
                             90: 'Quarter',
                             91: 'Quarter',
                             364 + leapadd: 'Annual',
                             365 + leapadd: 'Annual'}[delta.days]
            except KeyError:
                pass
        elif start.month == 6:
            if calendar.isleap(start.year + 1):
                leapadd = 1
            try:
                frequency = {29: 'Month',
                             30: 'Month',
                             91: 'Quarter',
                             92: 'Quarter',
                             364 + leapadd: 'Annual',
                             365 + leapadd: 'Annual'}[delta.days]
            except KeyError:
                pass
        elif start.month == 10:
            if calendar.isleap(start.year + 1):
                leapadd = 1
            try:
                frequency = {30: 'Month',
                             31: 'Month',
                             91: 'Quarter',
                             92: 'Quarter',
                             364 + leapadd: 'Annual',
                             365 + leapadd: 'Annual'}[delta.days]
            except KeyError:
                pass
        elif start.month == 11:
            if calendar.isleap(start.year + 1):
                leapadd = 1
                if start.day > 28:
                    quarteradd = 1
            try:
                frequency = {29: 'Month',
                             30: 'Month',
                             91 + quarteradd: 'Quarter',
                             92 + quarteradd: 'Quarter',
                             364 + leapadd: 'Annual',
                             365 + leapadd: 'Annual'}[delta.days]
            except KeyError:
                pass
        else:  # if start.month == 12:
            if calendar.isleap(start.year + 1):
                leapadd = 1
            try:
                frequency = {30: 'Month',
                             31: 'Month',
                             89 + leapadd: 'Quarter',
                             90 + leapadd: 'Quarter',
                             364 + leapadd: 'Annual',
                             365 + leapadd: 'Annual'}[delta.days]
            except KeyError:
                pass
    return frequency


def bsxml_lighting_system_lookup(src_type):
    lamp_type = et.Element('LampType')
    if src_type == 'CFL':
        et.SubElement(lamp_type, 'CompactFluorescent')
    elif src_type == 'Fluorescent T5/High output T5':  # Meh
        el = et.SubElement(lamp_type, 'LinearFluorescent')
        el = et.SubElement(el, 'LampLabel')
        el.text = 'T5'
    elif src_type == 'Fluorescent T8/Super T8':  # Meh
        el = et.SubElement(lamp_type, 'LinearFluorescent')
        el = et.SubElement(el, 'LampLabel')
        el.text = 'T8'
    elif src_type == 'Fluorescent T12/High output T12':  # Meh
        el = et.SubElement(lamp_type, 'LinearFluorescent')
        el = et.SubElement(el, 'LampLabel')
        el.text = 'T12'
    elif src_type == 'High pressure sodium':
        el = et.SubElement(lamp_type, 'HighIntensityDischarge')
        el = et.SubElement(el, 'LampLabel')
        el.text = 'Sodium Vapor High Pressure'
    elif src_type == 'Incandescent/Halogen':
        et.SubElement(lamp_type, 'Halogen')  # Meh
    elif src_type == 'Induction':
        et.SubElement(lamp_type, 'Induction')
    elif src_type == 'LED':
        el = et.SubElement(lamp_type, 'SolidStateLighting')
        el = et.SubElement(el, 'LampLabel')
        el.text = 'LED'
    elif src_type == 'Mercury vapor':
        el = et.SubElement(lamp_type, 'HighIntensityDischarge')
        el = et.SubElement(el, 'LampLabel')
        el.text = 'Mercury Vapor'
    elif src_type == 'Metal halide':
        el = et.SubElement(lamp_type, 'HighIntensityDischarge')
        el = et.SubElement(el, 'LampLabel')
        el.text = 'Metal Halide'
    else:  # src_type == 'Other':
        et.SubElement(lamp_type, 'Unknown')
    return lamp_type


def bsxml_lighting_control_lookup(control_type):
    control = None
    if control_type == 'Manual':
        control = et.Element('LightingControlTypeManual')
        control.text = 'Unknown'
    elif control_type == 'Occupancy sensor':
        control = et.Element('LightingControlTypeOccupancy')
        control.text = 'Unknown'
    elif control_type == 'Photocell':
        control = et.Element('LightingControlTypeDaylighting')  # Meh
        control.text = 'Unknown'
    elif control_type == 'Timer':
        control = et.Element('LightingControlTypeTimer')
        control.text = 'Unknown'
    # elif control_type == 'BAS':
    # elif control_type == 'Advanced':
    # elif control_type == 'Other':
    return control


def prettystring(element):
    return minidom.parseString(et.tostring(element, encoding='utf-8')).toprettyxml(indent='  ', encoding='utf-8')


def easymap(dictionary, inkey, outkey, parent, f=lambda x: x):
    if inkey in dictionary:
        if dictionary[inkey]:
            el = et.SubElement(parent, outkey)
            el.text = f(dictionary[inkey])


def easyremap(dictionary, inkey, outkey, parent, remap, f=lambda x: x):
    if inkey in dictionary:
        if dictionary[inkey]:
            el = et.SubElement(parent, outkey)
            el.text = f(remap[dictionary[inkey]])


def addel(outkey, parent, value):
    el = et.SubElement(parent, outkey)
    el.text = str(value)


def addudf(parent, key, value, create=True):
    udfs = parent.find('UserDefinedFields')
    if udfs is None:
        if not create:
            return
        udfs = et.SubElement(parent, 'UserDefinedFields')
    udf = et.SubElement(udfs, 'UserDefinedField')
    el = et.SubElement(udf, 'FieldName')
    el.text = key
    el = et.SubElement(udf, 'FieldValue')
    el.text = value


def appendudf(udfs, key, dictionary, prefix=None):
    if key in dictionary:
        udf = et.SubElement(udfs, 'UserDefinedField')
        el = et.SubElement(udf, 'FieldName')
        store_key = key
        if prefix:
            store_key = prefix + key
        el.text = store_key
        el = et.SubElement(udf, 'FieldValue')
        el.text = str(dictionary[key])


def easymapudf(dictionary, inkey, outkey, parent, f=lambda x: x):
    if inkey in dictionary and dictionary[inkey]:
        addudf(parent, outkey, f(dictionary[inkey]))


def yn2tf(s):
    return {'Y': 'true', 'N': 'false'}[s]


def repercentage(s):
    return str(s * 100) + '%'


def bsxml_condition_lookup(condition):
    return {'excellent': 'Excellent',
            'good': 'Good',
            'average': 'Average',
            'poor': 'Poor'}.get(condition.lower(), 'Other')


def bsxml_capacity_units_lookup(units):
    return {"cfh": "cfh",
            "ft3/min": "ft3/min",
            "cfm": "ft3/min",
            "kcf/h": "kcf/h",
            "mcf/day": "MCF/day",
            "gpm": "gpm",
            "w": "W",
            "kw": "kW",
            "hp": "hp",
            "mw": "MW",
            "btu/hr": "Btu/hr",
            "cal/h": "cal/h",
            "ft-lbf/h": "ft-lbf/h",
            "ft-lbf/min": "ft-lbf/min",
            "btu/s": "Btu/s",
            "kbtu/hr": "kBtu/hr",
            "mmbtu/hr": "MMBtu/hr",
            "therms/h": "therms/h",
            "lbs/h": "lbs/h",
            "klbs/h": "Klbs/h",
            "mlbs/h": "Mlbs/h",
            "cooling ton": "Cooling ton",
            "cooling tons": "Cooling ton",
            "tons": "Cooling ton",
            "ton": "Cooling ton"}.get(units.lower(), 'Other')


def map_equipment_inventory(inventory):
    hvacsystems = []
    heatrecoverysystems = []

    items = ['Description', 'Location', 'Type', 'Units', 'Rated efficiency (as applicable)',
             'Output Capacity', 'Area Served', 'Approx Year Installed',
             'Condition       (excellent, good, average, poor)']

    for name, data in inventory.items():
        if not 'Type' in data:
            # Could try something else here, but skip for now
            continue
        if data['Type'] == 'Heating Plant Type':
            system = et.Element('HVACSystem')
            system.attrib['ID'] = name
            el = et.SubElement(system, 'Plants')
            el = et.SubElement(el, 'HeatingPlant')
            easymap(data, 'Condition       (excellent, good, average, poor)', 'HeatingPlantCondition', el,
                    bsxml_condition_lookup)
            items = ['Description', 'Location', 'Units', 'Rated efficiency (as applicable)',
                     'Output Capacity', 'Area Served', 'Approx Year Installed']
            for item in items:
                easymapudf(data, item, 'ASHRAE Std 211 %s' % item, system, str)
            hvacsystems.append(system)
        elif data['Type'] == 'Cooling Plant Type':
            system = et.Element('HVACSystem')
            system.attrib['ID'] = name
            el = et.SubElement(system, 'Plants')
            el = et.SubElement(el, 'CoolingPlant')
            easymap(data, 'Condition       (excellent, good, average, poor)', 'CoolingPlantCondition', el,
                    bsxml_condition_lookup)
            items = ['Description', 'Location', 'Units', 'Rated efficiency (as applicable)',
                     'Output Capacity', 'Area Served', 'Approx Year Installed']
            for item in items:
                easymapudf(data, item, 'ASHRAE Std 211 %s' % item, system, str)
            hvacsystems.append(system)
        elif data['Type'] == 'Boiler Type':
            system = et.Element('HVACSystem')
            system.attrib['ID'] = name
            el = et.SubElement(system, 'Plants')
            hp = et.SubElement(el, 'HeatingPlant')
            el = et.SubElement(hp, 'Boiler')

            easymap(data, 'Output Capacity', 'OutputCapacity', el, str)
            easymap(data, 'Units', 'CapacityUnits', el, bsxml_capacity_units_lookup)
            easymap(data, 'Rated efficiency (as applicable)', 'ThermalEfficiency', el, str)

            easymap(data, 'Condition       (excellent, good, average, poor)', 'HeatingPlantCondition', hp,
                    bsxml_condition_lookup)
            items = ['Description', 'Location', 'Area Served', 'Approx Year Installed']
            for item in items:
                easymapudf(data, item, 'ASHRAE Std 211 %s' % item, system, str)
            hvacsystems.append(system)
        elif data['Type'] in ['Cooling Delivery Type', 'Heating Delivery Type']:
            typestring = data['Type'].replace(' Type', '')
            system = et.Element('HVACSystem')
            system.attrib['ID'] = name
            el = et.SubElement(system, 'HeatingAndCoolingSystems')
            el = et.SubElement(el, 'Delivery')

            easymap(data, 'Output Capacity', 'Capacity', el, str)
            easymap(data, 'Units', 'CapacityUnits', el, bsxml_capacity_units_lookup)
            easymap(data, 'Approx Year Installed', 'YearInstalled', el, str)

            addudf(system, 'ASHRAE Std 211 Type', typestring)

            items = ['Description', 'Location', 'Rated efficiency (as applicable)',
                     'Area Served', 'Condition       (excellent, good, average, poor)']
            for item in items:
                easymapudf(data, item, 'ASHRAE Std 211 %s' % item, system, str)
            hvacsystems.append(system)
        elif data['Type'] == 'Heat Recovery Type':
            system = et.Element('HeatRecoverySystem')
            system.attrib['ID'] = name

            easymap(data, 'Rated efficiency (as applicable)', 'HeatRecoveryEfficiency', el, str)
            easymap(data, 'Approx Year Installed', 'YearInstalled', el, str)

            items = ['Description', 'Location', 'Units', 'Output Capacity', 'Area Served',
                     'Condition       (excellent, good, average, poor)']
            for item in items:
                easymapudf(data, item, 'ASHRAE Std 211 %s' % item, system, str)
            heatrecoverysystems.append(system)
        else:  # elif data['Type'] == 'DX System Type':
            system = et.Element('HVACSystem')
            system.attrib['ID'] = name
            el = et.SubElement(system, 'HeatingAndCoolingSystems')
            cs = et.SubElement(el, 'CoolingSource')
            el = et.SubElement(cs, 'CoolingSourceType')
            et.SubElement(el, 'DX')

            easymap(data, 'Output Capacity', 'Capacity', cs, str)
            easymap(data, 'Units', 'CapacityUnits', cs, bsxml_capacity_units_lookup)
            easymap(data, 'Approx Year Installed', 'YearInstalled', cs, str)

            items = ['Description', 'Location', 'Rated efficiency (as applicable)',
                     'Area Served', 'Condition       (excellent, good, average, poor)']
            for item in items:
                easymapudf(data, item, 'ASHRAE Std 211 %s' % item, system, str)
            hvacsystems.append(system)
    return {'HVACSystem': hvacsystems, 'HeatRecoverySystem': heatrecoverysystems}


def map_to_buildingsync(obj, groupspaces=False):
    #
    allbuilding = obj['All - Building']
    spacefunctions = obj['All - Space Functions']
    metered_energy = obj['All - Metered Energy']
    delivered_energy = obj['All - Delivered Energy']
    summary = obj['L1 - EEM Summary']
    envelope = obj['L2 - Envelope']
    hvac = obj['L2 - HVAC']
    summary_L2 = obj['L2 - EEM Summary']
    lighting_plug_loads = obj['L2 - Lighting Elec & Plug Loads']
    inventory = obj['L2 - Equipment Inventory']
    #
    # All - Building
    #
    # Give the address
    address = et.Element('Address')
    if 'Street*' in allbuilding:
        el = et.SubElement(address, 'StreetAddressDetail')
        el = et.SubElement(el, 'Simplified')
        el = et.SubElement(el, 'StreetAddress')
        el.text = allbuilding['Street*']
    easymap(allbuilding, 'City*', 'City', address)
    easymap(allbuilding, 'State*', 'State', address)
    if 'Postal Code*' in allbuilding:
        postalcode = allbuilding['Postal Code*']
        postalcode, plus4 = process_zip(postalcode)
        postalcodeplus4 = postalcode
        if plus4:
            postalcodeplus4 += '-' + plus4
        el = et.SubElement(address, 'PostalCode')
        el.text = postalcode
        el = et.SubElement(address, 'PostalCodePlus4')
        el.text = postalcodeplus4
    # street address, city, state, zip5, zip5-4
    if len(address) == 0:
        address = None
    # Create contacts if they are present
    contacts = et.Element('Contacts')
    auditor = None
    if 'Energy Auditor' in allbuilding:
        auditor = et.SubElement(contacts, 'Contact')
        auditor.attrib['ID'] = 'EnergyAuditor'
        addel('ContactRole', auditor, 'Energy Auditor')
        addel('ContactName', auditor, allbuilding['Energy Auditor'])
    keycontact = None
    if 'Key Contact' in allbuilding:
        keycontact = et.SubElement(contacts, 'Contact')
        keycontact.attrib['ID'] = 'KeyContact'
        addel('ContactRole', keycontact, 'Other')
        addel('ContactName', keycontact, allbuilding['Key Contact'])
        addudf(keycontact, 'ASHRAE Standard 211 Role', 'Key Contact')
    if 'Client Name' in allbuilding:
        client = et.SubElement(contacts, 'Contact')
        client.attrib['ID'] = 'Client'
        addel('ContactRole', client, 'Other')
        addel('ContactName', client, allbuilding['Client Name'])
        addudf(client, 'ASHRAE Standard 211 Role', 'Client')
    if 'Building Owner' in allbuilding:
        owner = et.SubElement(contacts, 'Contact')
        owner.attrib['ID'] = 'BuildingOwner'
        addel('ContactRole', owner, 'Other')
        addel('ContactName', owner, allbuilding['Building Owner'])
        addudf(owner, 'ASHRAE Standard 211 Role', 'Owner')

    buildings = et.Element('Buildings')
    building = et.SubElement(buildings, 'Building')
    building.attrib['ID'] = 'Building'

    easymap(allbuilding, 'Building Name*', 'PremisesName', building)
    easymap(allbuilding, 'Building Description - Notable Conditions',
            'PremisesNotes', building)
    # OccupancyClassification should go here, but it can't: the enums don't match
    if 'Occupancy' in allbuilding:
        occupancy = allbuilding['Occupancy']
        if 'Typical number of occupants (during occ hours)' in occupancy:
            levels = et.SubElement(building, 'OccupancyLevels')
            level = et.SubElement(levels, 'OccupancyLevel')
            addel('OccupantQuantity', level,
                  str(occupancy['Typical number of occupants (during occ hours)']))
        typicalocc = et.Element('TypicalOccupantUsages')
        if 'Typical occupancy (hours/week)' in occupancy:
            occ = et.SubElement(typicalocc, 'TypicalOccupantUsage')
            addel('TypicalOccupantUsageValue', occ,
                  str(occupancy['Typical occupancy (hours/week)']))
            addel('TypicalOccupantUsageUnits', occ, 'Hours per week')
        if 'Typical occupancy (weeks/year)' in occupancy:
            occ = et.SubElement(typicalocc, 'TypicalOccupantUsage')
            addel('TypicalOccupantUsageValue', occ,
                  str(occupancy['Typical occupancy (weeks/year)']))
            addel('TypicalOccupantUsageUnits', occ, 'Weeks per year')
        if len(typicalocc) > 0:
            building.append(typicalocc)
        if 'Number of Dwelling Units in Building (Multifamily Only)' in occupancy:
            units = et.SubElement(building, 'SpatialUnits')
            addel('SpatialUnitType', units, 'Apartment units')
            addel('NumberOfUnits', units, str(occupancy['Number of Dwelling Units in Building (Multifamily Only)']))

    easymap(allbuilding, 'Conditioned Floors Above grade',
            'ConditionedFloorsAboveGrade', building, f=str)
    easymap(allbuilding, 'Conditioned Floors Below grade',
            'ConditionedFloorsBelowGrade', building, f=str)
    easymap(allbuilding, 'Building automation system? (Y/N)',
            'BuildingAutomationSystem', building, yn2tf)
    easymap(allbuilding, 'Historical landmark status? (Y/N)',
            'HistoricalLandmark', building, yn2tf)
    # Map to FloorAreas
    floorareas = et.Element('FloorAreas')
    if 'Total conditioned area' in allbuilding:
        floorarea = et.SubElement(floorareas, 'FloorArea')
        addel('FloorAreaType', floorarea, 'Conditioned')
        addel('FloorAreaValue', floorarea, allbuilding['Total conditioned area'])
    if 'Gross floor area' in allbuilding:
        floorarea = et.SubElement(floorareas, 'FloorArea')
        addel('FloorAreaType', floorarea, 'Gross')
        addel('FloorAreaValue', floorarea, allbuilding['Gross floor area'])
    if 'Conditioned area (heated only)' in allbuilding:
        floorarea = et.SubElement(floorareas, 'FloorArea')
        addel('FloorAreaType', floorarea, 'Cooled only')
        addel('FloorAreaValue', floorarea, allbuilding['Conditioned area (heated only)'])
    if 'Conditioned area (cooled only)' in allbuilding:
        floorarea = et.SubElement(floorareas, 'FloorArea')
        addel('FloorAreaType', floorarea, 'Heated only')
        addel('FloorAreaValue', floorarea, allbuilding['Conditioned area (cooled only)'])
    # Map Space Function table to FloorAreas
    if 'Space Function' in allbuilding:
        for key, value in allbuilding['Space Function'].items():
            floorarea = et.SubElement(floorareas, 'FloorArea')
            addel('FloorAreaType', floorarea, 'Custom')
            addel('FloorAreaCustomName', floorarea, key)
            addel('FloorAreaValue', floorarea, value)

    easymap(allbuilding, 'Year of construction*',
            'YearOfConstruction', building, f=str)

    easymap(allbuilding, 'Year of Prior Energy Audit',
            'YearOfLastEnergyAudit', building, f=str)

    easymap(allbuilding, 'Last Renovation*',
            'YearOfLastMajorRemodel', building, f=str)
    #
    # All - Space Functions
    #
    # subsections = et.Element('Subsections')
    spaces = []
    phvac = {}
    nohvac = []
    for key, value in spacefunctions.items():
        element = et.Element('Space')
        # First the stuff that has a slot to go into
        addel('PremisesName', element, key)
        if 'Number of Occupants' in value:
            levels = et.SubElement(element, 'OccupancyLevels')
            level = et.SubElement(levels, 'OccupancyLevel')
            addel('OccupantQuantity', level,
                  str(value['Number of Occupants']))
        typicalocc = et.Element('TypicalOccupantUsages')
        if 'Use (hours/week)' in value:
            occ = et.SubElement(typicalocc, 'TypicalOccupantUsage')
            addel('TypicalOccupantUsageValue', occ,
                  str(value['Use (hours/week)']))
            addel('TypicalOccupantUsageUnits', occ, 'Hours per week')
        if 'Use (weeks/year)' in value:
            occ = et.SubElement(typicalocc, 'TypicalOccupantUsage')
            addel('TypicalOccupantUsageValue', occ,
                  str(value['Use (weeks/year)']))
            addel('TypicalOccupantUsageUnits', occ, 'Weeks per year')
        if len(typicalocc) > 0:
            element.append(typicalocc)
        if 'Gross Floor Area' in value:
            floorareas = et.SubElement(element, 'FloorAreas')
            floorarea = et.SubElement(floorareas, 'FloorArea')
            addel('FloorAreaType', floorarea, 'Gross')
            addel('FloorAreaValue', floorarea, str(value['Gross Floor Area']))
        # Now for the UDFs
        easymapudf(value, 'Function type',
                   'ASHRAE Standard 211 Function Type', element)
        easymapudf(value, 'Original intended use',
                   'ASHRAE Standard 211 Original Intended Use', element)
        easymapudf(value, 'Percent Conditioned Area',
                   'ASHRAE Standard 211 Percent Conditioned Area', element,
                   f=repercentage)
        easymapudf(value, 'Approximate Plug Loads (W/sf)',
                   'ASHRAE Standard 211 Approximate Plug Loads', element, f=str)
        easymapudf(value, 'Principal HVAC Type',
                   'ASHRAE Standard 211 Principal HVAC Type', element, f=str)
        if value['Principal HVAC Type']:
            if value['Principal HVAC Type'] in phvac:
                phvac[value['Principal HVAC Type']].append(element)
            else:
                phvac[value['Principal HVAC Type']] = [element]
        else:
            nohvac.append(element)
        easymapudf(value, 'Principal Lighting Type',
                   'ASHRAE Standard 211 Principal Lighting Type', element, f=str)
        spaces.append(element)
    subsections = []
    subsection = None

    # Map the building shape if it is given
    if 'General Building Shape*' in envelope:
        subsections = et.SubElement(building, 'Subsections')
        subsection = et.SubElement(subsections, 'Subsection')
        addel('FootprintShape', subsection, envelope['General Building Shape*'])

    hvacsystems = None
    lightingsystems = None
    dhwsystems = None
    heatrecoverysystems = None
    wallsystems = None
    roofsystems = None
    ceilingsystems = None
    foundationsystems = None
    fenestrationsystems = None
    plugloads = None

    # L2 - HVAC, make one system to represent all of it.
    if len(hvac) > 0:
        hvacsystem = et.Element('HVACSystem')
        # Plant stuff
        if 'Boiler Type' in hvac:
            el = et.SubElement(hvacsystem, 'Plants')
            el = et.SubElement(el, 'HeatingPlant')
            el = et.SubElement(el, 'Boiler')
            for val in hvac['Boiler Type']:
                addudf(el, 'ASHRAE Std 211 Boiler Type', val)
        # HeatingAndCoolingSystems
        hacsys = el = et.Element('HeatingAndCoolingSystems')
        stuff = ['Heating Source', 'Heating Fuel']
        # Heating Source related info
        if any([el in hvac for el in stuff]):
            el = et.SubElement(hacsys, 'HeatingSource')
            for tag in stuff:
                if tag in hvac:
                    for val in hvac[tag]:
                        addudf(el, 'ASHRAE Std 211 %s' % tag, val)
        stuff = ['Cooling Source', 'Chiller Input', 'Compressor', 'Condenser']
        # Cooling Source related info
        if any([el in hvac for el in stuff]):
            el = et.SubElement(hacsys, 'CoolingSource')
            for tag in stuff:
                if tag in hvac:
                    for val in hvac[tag]:
                        addudf(el, 'ASHRAE Std 211 %s' % tag, val)
        if len(hacsys) > 0:
            hvacsystem.append(hacsys)

        # Tags with nowhere to go
        stuff = ['Zone Controls', 'Central Plant Controls', 'Heat Recovery', 'Outside Air',
                 'Cooling Distribution Equipment Type', 'Heating Distribution Equipment Type']
        for tag in stuff:
            if tag in hvac:
                for val in hvac[tag]:
                    addudf(hvacsystem, 'ASHRAE Std 211 %s' % tag, val)

        if len(hvacsystem) > 0:
            hvacsystem.attrib['ID'] = 'Std211L2HVAC'
            hvacsystems = et.Element('HVACSystems')
            hvacsystems.append(hvacsystem)

        stuff = ['SHW/DHW Source', 'SHW/DHW Fuel']
        if any([el in hvac for el in stuff]):
            dhwsystems = et.Element('DomesticHotWaterSystems')
            dhw = et.SubElement(dhwsystems, 'DomesticHotWaterSystem')
            dhw.attrib['ID'] = 'Std211L2HVACDHW'
            for tag in stuff:
                if tag in hvac:
                    for val in hvac[tag]:
                        addudf(dhw, 'ASHRAE Std 211 %s' % tag, val)

    if inventory:
        systems = map_equipment_inventory(inventory)
        if systems['HVACSystem']:
            if not hvacsystems:
                hvacsystems = et.Element('HVACSystems')
            for system in systems['HVACSystem']:
                hvacsystems.append(system)
        if systems['HeatRecoverySystem']:
            if not heatrecoverysystems:
                heatrecoverysystems = et.Element('HeatRecoverySystems')
            for system in systems['HeatRecoverySystem']:
                heatrecoverysystems.append(system)

    # Lighting
    if 'Lighting Source Type(s)' in lighting_plug_loads:
        num = 1
        sources = []
        for src_type, src in lighting_plug_loads['Lighting Source Type(s)'].items():
            source = et.Element('LightingSystem')
            source.attrib['ID'] = 'LightingSystem%d' % num
            num += 1
            source.append(bsxml_lighting_system_lookup(src_type))
            easyremap(src, 'Ballast Type(s)', 'BallastType', source, bsxml_ballast_lookup)
            control = bsxml_lighting_control_lookup(src['Control(s)'])
            if control == None:
                easymapudf(src, 'Control(s)', 'ASHRAE Std 211 Lighting Control', source)
            else:
                source.append(control)
            easymapudf(src, 'Space Type(s)*', 'ASHRAE Std 211 Space Type', source)
            easymapudf(src, 'Approx % Area Served', 'ASHRAE Std 211 Approx % Area Served', source, str)
            sources.append(source)
        if len(sources) > 0:
            lightingsystems = et.Element('LightingSystems')
            for src in sources:
                lightingsystems.append(src)

    # Plug/process loads
    if 'Major Process/Plug Load Type(s)**' in lighting_plug_loads:
        num = 1
        loads = []
        for ld_type, ld in lighting_plug_loads['Major Process/Plug Load Type(s)**'].items():
            load = et.Element('PlugLoad')
            addudf(load, 'ASHRAE Std 211 Major Process/Plug Load Type(s)', ld_type)
            easymapudf(ld, 'Key Operational Details***', 'ASHRAE Std 211 Key Operational Details', load)
            loads.append(load)
        if len(loads) > 0:
            plugloads = et.Element('PlugLoads')
            for load in loads:
                plugloads.append(load)

    # Handle sides
    if ('Total exposed above grade wall area (sq ft)' in envelope or
            'Total exposed above grade wall area R value' in envelope or
            'Glazing area, approx % of exposed wall area [10, 25, 50, 75, 90, 100]*' in envelope or
            'Wall Constructions' in envelope or
            'Fenestration Frame Types' in envelope or
            'Fenestration Glass Types' in envelope or
            'Fenestration Seal Condition' in envelope):
        # Something is there to put in sides, make what we need
        if subsection is None:
            subsections = et.SubElement(building, 'Subsections')
            subsection = et.SubElement(subsections, 'Subsection')
        sides = et.SubElement(subsection, 'Sides')
        side = et.SubElement(sides, 'Side')
        # Make a wall system if needed
        wallsystem = None
        if ('Total exposed above grade wall area (sq ft)' in envelope or
                'Total exposed above grade wall area R value' in envelope or
                'Glazing area, approx % of exposed wall area [10, 25, 50, 75, 90, 100]*' in envelope or
                'Wall Constructions' in envelope):
            wallsystems = et.Element('WallSystems')
            wallsystem = et.SubElement(wallsystems, 'WallSystem')
            wallsystem.attrib['ID'] = 'Wall1'
            easymap(envelope, 'Total exposed above grade wall area R value',
                    'WallRValue', wallsystem, f=str)
            easymapudf(envelope, 'Wall Constructions',
                       'ASHRAE Standard 211 Wall Construction', wallsystem, f=lambda x: ', '.join(x))
        # Make window stuff
        fenestrationsystem = None
        if ('Fenestration Frame Types' in envelope or
                'Fenestration Glass Types' in envelope):
            fenestrationsystems = et.Element('FenestrationSystems')
            fenestrationsystem = et.SubElement(fenestrationsystems, 'FenestrationSystem')
            fenestrationsystem.attrib['ID'] = 'Fenestration1'
            easymapudf(envelope, 'Fenestration Frame Types',
                       'ASHRAE Standard 211 Fenestration Frame Types',
                       fenestrationsystem, f=lambda x: ', '.join(x))
            easymapudf(envelope, 'Fenestration Glass Types',
                       'ASHRAE Standard 211 Fenestration Glass Types',
                       fenestrationsystem, f=lambda x: ', '.join(x))
            easymapudf(envelope, 'Fenestration Seal Condition',
                       'ASHRAE Standard 211 Fenestration Seal Condition',
                       fenestrationsystem)
            easymapudf(envelope, 'Description of Exterior doors**',
                       'ASHRAE Standard 211 Description of Exterior doors',
                       fenestrationsystem)
        # Fill in the side information
        if wallsystem is not None:
            wallid = et.SubElement(side, 'WallID')
            wallid.attrib['IDref'] = wallsystem.attrib['ID']
            if 'Total exposed above grade wall area (sq ft)' in envelope:
                addel('WallArea', wallid,
                      str(envelope['Total exposed above grade wall area (sq ft)']))
        if fenestrationsystem is not None:
            windowid = et.SubElement(side, 'WindowID')
            windowid.attrib['IDref'] = fenestrationsystem.attrib['ID']
            if 'Glazing area, approx % of exposed wall area [10, 25, 50, 75, 90, 100]*' in envelope:
                addel('WindowToWallRatio', windowid,
                      str(envelope['Glazing area, approx % of exposed wall area [10, 25, 50, 75, 90, 100]*']))
    # Roof is next
    if ('Roof area (sq ft)' in envelope or
            'Roof area R value' in envelope or
            'Cool Roof (Y/N)' in envelope or
            'Roof condition' in envelope or
            'Roof Construction' in envelope):
        roofsystems = et.Element('RoofSystems')
        roofsystem = et.SubElement(roofsystems, 'RoofSystem')
        roofsystem.attrib['ID'] = 'Roof1'
        easymap(envelope, 'Roof area R value', 'RoofRValue',
                roofsystem, f=str)
        easymapudf(envelope, 'Cool Roof (Y/N)',
                   'ASHRAE Standard 211 Cool Roof (Y/N)', roofsystem)
        easymapudf(envelope, 'Roof condition',
                   'ASHRAE Standard 211 Roof Condition', roofsystem)
        easymapudf(envelope, 'Roof Construction',
                   'ASHRAE Standard 211 Roof Construction',
                   roofsystem, f=lambda x: ', '.join(x))
        roofid = et.SubElement(subsection, 'RoofID')
        roofid.attrib['IDref'] = roofsystem.attrib['ID']
        easymap(envelope, 'Roof area (sq ft)', 'RoofArea', roofid, f=str)

    # Make a ceiling system if needed
    if 'Floor Construction' in envelope:
        if ('Steel joist' in envelope['Floor Construction'] or
                'Wood frame' in envelope['Floor Construction']):
            value = []
            if 'Steel joist' in envelope['Floor Construction']:
                value = ['Steel joist']
            if 'Wood frame' in envelope['Floor Construction']:
                value.append('Wood frame')
            value = ', '.join(value)
            ceilingsystems = et.Element('CeilingSystems')
            ceilingsystem = et.SubElement(ceilingsystems, 'CeilingSystem')
            ceilingsystem.attrib['ID'] = 'Ceiling1'
            addudf(ceilingsystem, 'ASHRAE Standard 211 Floor Construction',
                   str(value))
            ceilingid = et.SubElement(subsection, 'CeilingID')
            ceilingid.attrib['IDref'] = ceilingsystem.attrib['ID']

    # Foundation systems
    foundationsystem = None
    if ('Foundation Type' in envelope or
            'Floor Construction' in envelope):
        foundationsystems = et.Element('FoundationSystems')
        foundationsystem = et.SubElement(foundationsystems, 'FoundationSystem')
        foundationsystem.attrib['ID'] = 'Foundation1'
        easymapudf(envelope, 'Foundation Type',
                   'ASHRAE Standard 211 Foundation Type',
                   foundationsystem, f=lambda x: ', '.join(x))
        easymapudf(envelope, 'Floor Construction',
                   'ASHRAE Standard 211 Floor Construction',
                   foundationsystem, f=lambda x: ', '.join(x))
        foundationid = et.SubElement(subsection, 'FoundationID')
        foundationid.attrib['IDref'] = foundationsystem.attrib['ID']

    # Map the UDFs from L2 - Envelope
    udfs = et.Element('UserDefinedFields')
    appendudf(udfs, 'Below grade wall area (sq ft)', envelope, prefix='ASHRAE Standard 211 ')
    appendudf(udfs, 'Below grade wall area (sq m)', envelope, prefix='ASHRAE Standard 211 ')
    appendudf(udfs, 'Overall Enclosure Tightness Assessment', envelope, prefix='ASHRAE Standard 211 ')
    appendudf(udfs, 'Description of Exterior doors**', envelope, prefix='ASHRAE Standard 211 ')
    appendudf(udfs, 'Below grade wall area R value', envelope, prefix='ASHRAE Standard 211 ')
    appendudf(udfs, 'Above grade wall common area with other conditioned buildings (ft2)', envelope,
              prefix='ASHRAE Standard 211 ')
    appendudf(udfs, 'Above grade wall common area with other conditioned buildings (m2)', envelope,
              prefix='ASHRAE Standard 211 ')
    # appendudf(udfs, 'Fenestration Seal Condition', envelope, prefix = 'ASHRAE Standard 211 ')

    if len(udfs) > 0:
        if subsection is None:
            subsections = et.SubElement(building, 'Subsections')
            subsection = et.SubElement(subsections, 'Subsection')
        subsection.append(udfs)

    thermalzones = []
    if len(spaces) > 0:
        if groupspaces:
            # Group spaces by the principle HVAC type
            thermalzones = et.Element('ThermalZones')
            for phvactype, spcs in phvac.items():
                tz = et.SubElement(thermalzones, 'ThermalZone')
                tzspaces = et.SubElement(tz, 'Spaces')
                for space in spcs:
                    tzspaces.append(space)
            # Anything with nothing gets its own zone
            for space in nohvac:
                tz = et.Element('ThermalZone')
                tzspaces = et.SubElement(tz, 'Spaces')
                tzspaces.append(space)
        else:
            # Every space gets its own thermal zone
            thermalzones = et.Element('ThermalZones')
            for space in spaces:
                tz = et.SubElement(thermalzones, 'ThermalZone')
                tzspaces = et.SubElement(tz, 'Spaces')
                tzspaces.append(space)
    if len(thermalzones) > 0:
        if subsection is None:
            subsections = et.SubElement(building, 'Subsections')
            subsection = et.SubElement(subsections, 'Subsection')
        subsection.append(thermalzones)

    # Now for the UDFs from All - Building
    easymapudf(allbuilding, 'Primary Building use type*',
               'ASHRAE Standard 211 Primary Building Use Type', building)
    easymapudf(allbuilding, 'Year Last Commissioned',
               'ASHRAE Standard 211 Year Last Commissioned', building, f=str)
    easymapudf(allbuilding, 'Percent owned (%)',
               'ASHRAE Standard 211 Percent Owned', building, f=repercentage)
    easymapudf(allbuilding, 'Percent leased (%)',
               'ASHRAE Standard 211 Percent Leased', building, f=repercentage)
    easymapudf(allbuilding, 'Total Number of Floors',
               'ASHRAE Standard 211 Total Number of Floors', building, f=str)
    if 'Excluded Spaces' in allbuilding:
        allbuilding['Excluded Spaces'] = ', '.join(allbuilding['Excluded Spaces'])
    easymapudf(allbuilding, 'Excluded Spaces',
               'ASHRAE Standard 211 Excluded Spaces', building)

    if 'Occupancy' in allbuilding:
        easymapudf(allbuilding['Occupancy'],
                   '% of Dwelling Units currently Occupied (Multifamily Only)',
                   'ASHRAE Standard 211 Percent Dwelling Units Currently Occupied',
                   building, f=repercentage)

    # Wrap up for building
    if len(building) == 0:
        building = None
        buildings = None

    # Map energy sources, metered energy, and delivered energy to a report
    report = et.Element('Report')
    scenario = None
    resources = None

    if ('Energy Sources' in allbuilding
            or 'Utility #1' in metered_energy
            or 'Utility #2' in metered_energy
            or 'Utility #3' in metered_energy
            or delivered_energy != {}):
        scenarios = et.SubElement(report, 'Scenarios')
        scenario = et.SubElement(scenarios, 'Scenario')
        scenario.attrib['ID'] = 'ASHRAEStandard211Scenario'
        addel('ScenarioName', scenario, 'ASHRAE Standard 211 Scenario')
        resources = et.SubElement(scenario, 'ResourceUses')

    #
    # Map the energy sources from 'All - Building', does this need to be
    # harmonized with the information from 'All - Metered Energy' below?
    #
    if 'Energy Sources' in allbuilding:
        for el in allbuilding['Energy Sources']:
            resource = et.Element('ResourceUse')
            # Nope, enum fail on both
            # easymap(el, 'Energy Source', 'EnergyResource', resource)
            # if 'Type' in el:
            #    sub = et.SubElement(resource, 'Utility')
            #    sub = et.SubElement(sub, 'MeteringConfiguration')
            #    sub.text = el['Type']
            easymapudf(el, 'Energy Source', 'ASHRAE Standard 211 Energy Source',
                       resource)
            easymapudf(el, 'Type', 'ASHRAE Standard 211 Type', resource)
            easymapudf(el, 'ID', 'ASHRAE Standard 211 ID', resource, f=str)
            easymapudf(el, 'Rate schedule', 'ASHRAE Standard 211 Rate Schedule',
                       resource, f=str)
            if len(resource) > 0:
                resources.append(resource)

    # Add resource uses for metered and delivered energy
    for name in ['Utility #1', 'Utility #2', 'Utility #3']:
        if name in metered_energy:
            resource = et.Element('ResourceUse')
            resource.attrib['ID'] = 'Std211ResourceUse' + name.replace(' #', '')
            if metered_energy[name]['Definition']['Units'].startswith("=INDEX('Drop Down Lists'!"):
                # Use default
                metered_energy[name]['Definition']['Units'] = metered_energy_default_units[metered_energy[name]['Type']]
            if metered_energy[name]['Definition']['kBtu/unit'].startswith('=IFERROR(INDEX(EnergyConversionRates,MATCH'):
                # Use default
                metered_energy[name]['Definition']['kBtu/unit'] = str(
                    conversion_to_kBtu[metered_energy[name]['Definition']['Units']])
            if metered_energy[name]['Type'] in metered_energy_type_lookup:
                el = et.SubElement(resource, 'EnergyResource')
                el.text = metered_energy_type_lookup[metered_energy[name]['Type']]
            else:
                el = et.SubElement(resource, 'EnergyResource')
                el.text = 'Other'
                easymapudf(metered_energy[name], 'Type',
                           'ASHRAE Standard 211 Energy Source', resource)
            el = et.SubElement(resource, 'ResourceUnits')
            el.text = metered_energy_bsxml_units[metered_energy[name]['Type']]
            el = et.SubElement(resource, 'UtilityID')
            el.attrib['IDref'] = 'Std211Metered' + name.replace(' #', '')
            easymapudf(metered_energy[name]['Definition'], 'kBtu/unit', 'ASHRAE Standard 211 kBtu/unit', resource)
            resources.append(resource)

    if delivered_energy:
        resource = et.Element('ResourceUse')
        resource.attrib['ID'] = 'Std211ResourceUseDelivered1'
        if delivered_energy['Definition']['Conversion to kBTU'].startswith("=IFERROR(INDEX("):
            # Use default
            delivered_energy['Definition']['Conversion to kBTU'] = str(
                conversion_to_kBtu[delivered_energy['Definition']['Units']])
        el = et.SubElement(resource, 'EnergyResource')
        fueltype = delivered_energy['Definition']['Delivered Energy Type (if applicable)']
        if fueltype == 'Oil':
            fueltype = 'Fuel oil'
        el.text = fueltype
        el = et.SubElement(resource, 'ResourceUnits')
        el.text = bsxml_unit_lookup[delivered_energy['Definition']['Units']]
        easymapudf(delivered_energy['Definition'], 'Conversion to kBTU', 'ASHRAE Standard 211 Conversion to kBTU',
                   resource)
        if 'Estimated Annual Use**' in delivered_energy['Definition']:
            easymapudf(delivered_energy['Definition'], 'Estimated Annual Use**',
                       'ASHRAE Standard 211 Estimated Annual Use', resource,
                       str)
        resources.append(resource)

    # Now the time series data
    datapoints = []

    keys = {'Utility #1': {'Use': 'Energy', 'Cost': 'Currency', 'Peak': 'Energy'},
            'Utility #2': {'Use': 'Energy', 'Cost': 'Currency'},
            'Utility #3': {'Use': 'Energy', 'Cost': 'Currency'}}

    reading_type = {'Use': 'Total',
                    'Cost': 'Total',
                    'Peak': 'Peak'}

    for name in ['Utility #1', 'Utility #2', 'Utility #3']:
        if name in metered_energy:
            refname = 'Std211ResourceUse' + name.replace(' #', '')
            if 'Data' in metered_energy[name]:
                for pt in metered_energy[name]['Data']:
                    start = pt['Start Date']
                    end = pt['End Date']
                    # Compute the frequency, we don't handle 'Unknown'
                    frequency = determine_frequency(start, end)
                    for inkey, outkey in keys[name].items():
                        ts = et.Element('TimeSeries')
                        el = et.SubElement(ts, 'ReadingType')
                        el.text = reading_type[inkey]
                        el = et.SubElement(ts, 'TimeSeriesReadingQuantity')
                        el.text = outkey
                        el = et.SubElement(ts, 'StartTimeStamp')
                        el.text = start.strftime('%Y-%m-%dT00:00:00')
                        el = et.SubElement(ts, 'EndTimeStamp')
                        el.text = end.strftime('%Y-%m-%dT00:00:00')
                        el = et.SubElement(ts, 'IntervalFrequency')
                        el.text = frequency
                        el = et.SubElement(ts, 'IntervalReading')
                        el.text = str(pt[inkey])
                        el = et.SubElement(ts, 'ResourceUseID')
                        el.attrib['IDref'] = refname
                        datapoints.append(ts)

    if delivered_energy:
        refname = 'Std211ResourceUseDelivered1'
        if 'Data' in delivered_energy:
            for pt in delivered_energy['Data']:
                start = pt['Delivery date']
                for inkey, outkey in {'Volume': 'Other', 'Cost': 'Currency'}.items():
                    ts = et.Element('TimeSeries')
                    el = et.SubElement(ts, 'ReadingType')
                    el.text = 'Total'
                    el = et.SubElement(ts, 'TimeSeriesReadingQuantity')
                    el.text = outkey
                    el = et.SubElement(ts, 'StartTimeStamp')
                    el.text = start.strftime('%Y-%m-%dT00:00:00')
                    el = et.SubElement(ts, 'IntervalReading')
                    el.text = str(pt[inkey])
                    el = et.SubElement(ts, 'ResourceUseID')
                    el.attrib['IDref'] = refname
                    datapoints.append(ts)

    if len(datapoints) > 0:
        ts = et.SubElement(scenario, 'TimeSeriesData')
        for pt in datapoints:
            ts.append(pt)

    if len(scenario) > 0 and (building is not None):
        link = et.SubElement(scenario, 'LinkedPremises')
        el = et.SubElement(link, 'Building')
        el = et.SubElement(el, 'LinkedBuildingID')
        el.attrib['IDref'] = building.attrib['ID']

    # Add the utility items
    utilities = et.Element('Utilities')
    for name in ['Utility #1', 'Utility #2', 'Utility #3']:
        if name in metered_energy:
            el = et.SubElement(utilities, 'Utility')
            el.attrib['ID'] = 'Std211Metered' + name.replace(' #', '')
            el = et.SubElement(el, 'UtilityName')
            el.text = name
    if len(utilities) > 0:
        report.append(utilities)

    if auditor is not None:
        el = et.SubElement(report, 'AuditorContactID')
        el.attrib['IDref'] = auditor.attrib['ID']

    easymapudf(allbuilding, 'Date of site visit(s)',
               'ASHRAE Standard 211 Date of site visit(s)', report)

    # Wrap up for report
    if len(report) == 0:
        report = None
    #
    # L1 - EEM Summary
    #
    fields = ['Modified System',
              'Impact on Occupant Comfort or IEQ',
              'Other Non-Energy Impacts',
              'Cost',
              'Savings Impact',
              'Typical ROI',
              'Priority']
    # First the low cost items
    measures = et.Element('Measures')
    if 'Low-Cost and No-Cost Recommendations' in summary:
        for key, value in summary['Low-Cost and No-Cost Recommendations'].items():
            measure = et.SubElement(measures, 'Measure')
            el = et.SubElement(measure, 'LongDescription')
            el.text = key
            udfs = et.SubElement(measure, 'UserDefinedFields')
            for field in fields:
                if field in value:
                    udf = et.SubElement(udfs, 'UserDefinedField')
                    udfname = et.SubElement(udf, 'FieldName')
                    udfname.text = field
                    udfvalue = et.SubElement(udf, 'FieldValue')
                    udfvalue.text = value[field]
            udf = et.SubElement(udfs, 'UserDefinedField')
            udfname = et.SubElement(udf, 'FieldName')
            udfname.text = 'ASHRAE Standard 211 L1 Measure Category'
            udfvalue = et.SubElement(udf, 'FieldValue')
            udfvalue.text = 'Low-Cost and No-Cost Recommendations'
    # Change that one thing...
    fields[1] = 'Impact on Occupant Comfort'
    if 'Potential Capital Recommendations' in summary:
        for key, value in summary['Potential Capital Recommendations'].items():
            measure = et.SubElement(measures, 'Measure')
            el = et.SubElement(measure, 'LongDescription')
            el.text = key
            udfs = et.SubElement(measure, 'UserDefinedFields')
            for field in fields:
                if field in value:
                    udf = et.SubElement(udfs, 'UserDefinedField')
                    udfname = et.SubElement(udf, 'FieldName')
                    udfname.text = field
                    udfvalue = et.SubElement(udf, 'FieldValue')
                    udfvalue.text = value[field]
            udf = et.SubElement(udfs, 'UserDefinedField')
            udfname = et.SubElement(udf, 'FieldName')
            udfname.text = 'ASHRAE Standard 211 L2 Measure Category'
            udfvalue = et.SubElement(udf, 'FieldValue')
            udfvalue.text = 'Potential Capital Recommendations'

    #
    # L2 - EEM Summary
    #
    udf_fields = ['Electricity Cost Savings', 'Non-energy Cost Savings']
    # Try to build the utility savings headings
    utility_units = []
    utility_types = []
    for name in ['Utility #1', 'Utility #2', 'Utility #3']:
        if name in metered_energy:
            utility_units.append(metered_energy[name]['Definition']['Units'])
            utility_types.append(metered_energy[name]['Type'])
    if delivered_energy:
        utility_types.append(delivered_energy['Definition']['Delivered Energy Type (if applicable)'])
        utility_units.append(delivered_energy['Definition']['Units'])
    for category, eems in summary_L2.items():
        for key, value in eems.items():
            measure = et.SubElement(measures, 'Measure')
            el = et.SubElement(measure, 'LongDescription')
            el.text = key
            measure_savings = et.Element('MeasureSavingsAnalysis')

            annual_by_fuels = et.Element('AnnualSavingsByFuels')
            utilnum = 1
            for util_units, util_type in zip(utility_units, utility_types):
                if utilnum == 4:
                    header = 'Delivered Energy'
                else:
                    header = 'Utility #%d' % utilnum  # util_type + ' [' + util_units +']'
                utilnum += 1
                if header in value:
                    if value[header]:
                        savings = et.SubElement(annual_by_fuels, 'AnnualSavingsByFuel')
                        el = et.SubElement(savings, 'EnergyResource')
                        el.text = metered_energy_type_lookup[util_type]
                        el = et.SubElement(savings, 'ResourceUnits')
                        el.text = bsxml_unit_lookup[util_units]
                        el = et.SubElement(savings, 'AnnualSavingsNativeUnits')
                        el.text = str(value[header])

            if len(annual_by_fuels) > 0:
                measure_savings.append(annual_by_fuels)

            easymap(value, 'Potential Incentives', 'FundingFromIncentives', measure_savings, str)

            if len(measure_savings) > 0:
                measure.append(measure_savings)

            easymap(value, 'Measure Life (years)', 'UsefulLife', measure, str)
            easymap(value, 'Measure Cost', 'MeasureTotalFirstCost', measure, str)

            udfs = et.SubElement(measure, 'UserDefinedFields')
            for field in udf_fields:
                if field in value:
                    if value[field]:
                        udf = et.SubElement(udfs, 'UserDefinedField')
                        udfname = et.SubElement(udf, 'FieldName')
                        udfname.text = 'ASHRAE Std 211 ' + field
                        udfvalue = et.SubElement(udf, 'FieldValue')
                        udfvalue.text = value[field]
            udf = et.SubElement(udfs, 'UserDefinedField')
            udfname = et.SubElement(udf, 'FieldName')
            udfname.text = 'ASHRAE Standard 211 L2 Measure Category'
            udfvalue = et.SubElement(udf, 'FieldValue')
            udfvalue.text = category

    #
    # Assemble the final result
    #
    attr_qname = et.QName("http://www.w3.org/2001/XMLSchema-instance", "schemaLocation")
    nsmap = {None: "http://buildingsync.net/schemas/bedes-auc/2019",
             'xsi': "http://www.w3.org/2001/XMLSchema-instance"}
    bsxml = et.Element('BuildingSync',
                       {attr_qname: "http://buildingsync.net/schemas/bedes-auc/2019 ../BuildingSync.xsd"},
                       nsmap=nsmap)
    # The following five lines are the original ElementTree version
    # bsxml = et.Element('Audits')
    # bsxml.attrib['xmlns'] = "http://nrel.gov/schemas/bedes-auc/2014"
    # bsxml.attrib['xmlns:xsi'] = "http://www.w3.org/2001/XMLSchema-instance"
    # bsxml.attrib['xsi:schemaLocation'] = "http://nrel.gov/schemas/bedes-auc/2014 ../BuildingSync.xsd"

    # First is Sites
    facilities = None
    if (address is not None) or (keycontact is not None) or (buildings is not None):
        facilities = et.SubElement(bsxml, 'Facilities')
        facility = et.SubElement(facilities, 'Facility')
        sites = et.SubElement(facility, 'Sites')
        site = et.SubElement(sites, 'Site')
        if address is not None:
            site.append(address)
        if keycontact is not None:
            pcid = et.SubElement(site, 'PrimaryContactID')
            pcid.text = keycontact.attrib['ID']
        if buildings is not None:
            site.append(buildings)
    # Second is Systems
    if ((hvacsystems is not None) or (lightingsystems is not None) or (dhwsystems is not None)
            or (heatrecoverysystems is not None) or (wallsystems is not None) or (roofsystems is not None)
            or (ceilingsystems is not None) or (fenestrationsystems is not None) or (foundationsystems is not None)
            or (plugloads is not None)):
        if facilities is None:
            facilities = et.SubElement(bsxml, 'Facilities')
            facility = et.SubElement(facilities, 'Facility')
        systems = et.SubElement(facility, 'Systems')
        if hvacsystems is not None:
            systems.append(hvacsystems)
        if lightingsystems is not None:
            systems.append(lightingsystems)
        if dhwsystems is not None:
            systems.append(dhwsystems)
        if heatrecoverysystems is not None:
            systems.append(heatrecoverysystems)
        if wallsystems is not None:
            systems.append(wallsystems)
        if roofsystems is not None:
            systems.append(roofsystems)
        if ceilingsystems is not None:
            systems.append(ceilingsystems)
        if fenestrationsystems is not None:
            systems.append(fenestrationsystems)
        if foundationsystems is not None:
            systems.append(foundationsystems)
        if plugloads is not None:
            systems.append(plugloads)
    # Next is Measures
    if measures is not None:
        if facilities is None:
            facilities = et.SubElement(bsxml, 'Facilities')
            facility = et.SubElement(facilities, 'Facility')
        facility.append(measures)
    # Now Reports
    if report is not None:
        if facilities is None:
            facilities = et.SubElement(bsxml, 'Facilities')
            facility = et.SubElement(facilities, 'Facility')
        facility.append(report)
    # Last is Contacts
    if contacts is not None:
        if facilities is None:
            facilities = et.SubElement(bsxml, 'Facilities')
            facility = et.SubElement(facilities, 'Facility')
        facility.append(contacts)
    # Done!  
    return bsxml


def map_std211_xlsx_to_string(filename, verbose=False, groupspaces=False):
    if not os.path.exists(filename):
        raise Exception('File "%s" does not exist' % filename)
    if verbose:
        wb = loadxl.load_workbook(filename)
    else:
        warnings.simplefilter("ignore")
        wb = loadxl.load_workbook(filename)
        warnings.simplefilter("default")
    std211 = read_std211_xlsx(wb)
    bsxml = map_to_buildingsync(std211, groupspaces=groupspaces)
    return '<?xml version="1.0" encoding="UTF-8"?>' + et.tostring(bsxml, encoding='utf-8').decode('utf-8')


def map_std211_xlsx_to_prettystring(filename, verbose=False, groupspaces=False):
    if not os.path.exists(filename):
        raise Exception('File "%s" does not exist' % filename)
    if verbose:
        wb = loadxl.load_workbook(filename)
    else:
        warnings.simplefilter("ignore")
        wb = loadxl.load_workbook(filename)
        warnings.simplefilter("default")
    std211 = read_std211_xlsx(wb)
    bsxml = map_to_buildingsync(std211, groupspaces=groupspaces)
    return prettystring(bsxml).decode('utf-8')


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='Convert ASHRAE Std. 211 Workbook into BSXML.')
    parser.add_argument('infile', metavar='INFILE')
    parser.add_argument('-p', '--pretty', dest='pretty', action='store_true',
                        help='output pretty xml')
    parser.add_argument('-o', '--output', dest='outfile', action='store',
                        default='std211.xml',
                        help='file to save BSXML output in')
    parser.add_argument('-g', '--groupspaces', dest='group', action='store_true',
                        help='group spaces into zones by principal HVAC type')
    parser.add_argument('-v', '--verbose', dest='verbose', action='store_true',
                        help='operate verbosely')

    args = parser.parse_args()

    if not os.path.exists(args.infile):
        raise Exception('File "%s" does not exist' % args.infile)

    if args.verbose:
        wb = loadxl.load_workbook(args.infile)
    else:
        warnings.simplefilter("ignore")
        wb = loadxl.load_workbook(args.infile)
        warnings.simplefilter("default")

    std211 = read_std211_xlsx(wb)
    bsxml = map_to_buildingsync(std211, groupspaces=args.group)
    if args.verbose:
        print(prettystring(bsxml).decode('utf-8'))
    fp = open(args.outfile, 'w')
    if args.pretty:
        fp.write(prettystring(bsxml).decode('utf-8'))
    else:
        fp.write('<?xml version="1.0" encoding="UTF-8"?>')
        fp.write(et.tostring(bsxml, encoding='utf-8').decode('utf-8'))
    fp.close()
