# BuildingSync(R), Copyright (c) 2015-2020, Alliance for Sustainable Energy, LLC.
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without modification,
# are permitted provided that the following conditions are met:
#
# (1) Redistributions of source code must retain the above copyright notice, this
#     list of conditions and the following disclaimer.
#
# (2) Redistributions in binary form must reproduce the above copyright notice,
#     this list of conditions and the following disclaimer in the documentation
#     and/or other materials provided with the distribution.
#
# (3) Neither the name of the copyright holder nor the names of any contributors
#     may be used to endorse or promote products derived from this software
#     without specific prior written permission from the respective party.
#
# (4) Other than as required in clauses (1) and (2), distributions in any form of
#     modifications or other derivative works may not use the "BuildingSync"
#     trademark or any other confusingly similar designation without specific
#     prior written permission from Alliance for Sustainable Energy, LLC.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDER(S) AND ANY CONTRIBUTORS "AS
# IS" AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
# IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
# DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER(S), ANY CONTRIBUTORS, THE
# UNITED STATES GOVERNMENT, OR THE UNITED STATES DEPARTMENT OF ENERGY, NOR ANY
# OF THEIR EMPLOYEES, BE LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL,
# EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO,
# PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR
# BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER
# IN CONTRACT, STRICT LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE)
# ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE
# POSSIBILITY OF SUCH DAMAGE.

import openpyxl
import zipfile
import os
import xml.etree.ElementTree as et

class Control:
    """Data structure to contain data associated with an Excel control

    :ivar text: Text in the control
    :ivar checked: Boolean checkbox status, True if checked and False otherwise

    """
    def __init__(self, name, relId=None, shapeId=None):
        self.name = name # name of the control
        self.relId = relId
        self.shapeId = shapeId
        self.text = None
        self.checked = False
        self._ctrlProp = None

ns = {'s':"http://schemas.openxmlformats.org/spreadsheetml/2006/main",
      'r':"http://schemas.openxmlformats.org/officeDocument/2006/relationships",
      'ns0':"http://schemas.openxmlformats.org/package/2006/relationships",
      'xdr':"http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
      'a':"http://schemas.openxmlformats.org/drawingml/2006/main"}

def normpath(path):
    npath = os.path.normpath(path)
    if os.sep == '\\':
        npath = npath.replace('\\','/')
    return npath

def load_workbook(filename, control_sheets=None):
    """Load an Excel spreadsheet into memory including controls and textboxes

    :param filename: file name of Excel file read
    :param control_sheets: if not none, the list of names of spreadsheets to process. If not specified, all sheets will be processed
    :return: openpyxl workbook object with appended controls and textboxes
    """
    workbook = openpyxl.load_workbook(filename) #,read_only=True,keep_vba=True)
    if not control_sheets:
        control_sheets = workbook.sheetnames
    sheets = []
    for number,name in enumerate(workbook.sheetnames):
        if name in control_sheets:
            sheets.append((number+1,name))
    archive = zipfile.ZipFile(filename) #workbook._archive
    for number,sheetname in sheets:
        sheet = workbook[sheetname]
        sheet.controls = {}
        sheet.textboxes = {}
        # Find and read the sheet
        sheetxml = 'sheet%d.xml' % number
        for path in archive.namelist():
            if path.endswith(sheetxml):
                break
        else:
            continue
        xmltxt = archive.read(path)
        # Read the controls in the sheet
        controlxml = et.fromstring(xmltxt).findall('.//s:control', ns)
        for control in controlxml:
            obj = Control(control.attrib['name'],
                          shapeId=control.attrib['shapeId'],
                          relId = control.attrib['{%s}id'%ns['r']])
            sheet.controls[obj.name] = obj
        # Find and read the relationships sheet
        sheetxml = 'sheet%d.xml.rels' % number
        for path in archive.namelist():
            if path.endswith(sheetxml):
                break
        else:
            continue
        # Get the relationship for the sheet itself
        drawingxml = et.fromstring(xmltxt).findall('.//s:drawing', ns)
        if len(drawingxml) != 1:
            continue
        drawingId = drawingxml[0].attrib['{%s}id'%ns['r']]
        # Now for the controls
        xmltxt = archive.read(path)
        rels = et.fromstring(xmltxt).findall('.//ns0:Relationship', ns)
        drawingfile = None
        for rel in rels:
            if rel.attrib['Id'] == drawingId:
                drawingfile = rel.attrib['Target']
                if drawingfile.startswith('..'):
                    drawingfile = '../'+drawingfile
                drawingfile = normpath(path.replace(sheetxml,drawingfile))
        # Mine the drawing file for the names of the controls
        if drawingfile:
            xmltxt = archive.read(drawingfile)
            drawing = et.fromstring(xmltxt)
            anchors = drawing.findall('.//xdr:absoluteCellAnchor', ns)
            anchors.extend(drawing.findall('.//xdr:twoCellAnchor', ns))
            anchors.extend(drawing.findall('.//xdr:oneCellAnchor', ns))
            for anchor in anchors:
                t = anchor.findall('.//a:t', ns)
                if not t:
                    continue
                cnvpr = anchor.findall('.//xdr:cNvPr', ns)[0]
                try:
                    sheet.controls[cnvpr.attrib['name']].text = t[0].text
                except KeyError:
                    sheet.textboxes[cnvpr.attrib['name']] = t[0].text
        # Find and get info from the individual property files
        for name,control in sheet.controls.items():
            for rel in rels:
                if rel.attrib['Id'] == control.relId:
                    propfile = rel.attrib['Target']
                    if propfile.startswith('..'):
                        propfile = '../'+propfile
                    propfile = normpath(path.replace(sheetxml,propfile))
                    xmltxt = archive.read(propfile)
                    form = et.fromstring(xmltxt)
                    if 'checked' in form.attrib:
                        if form.attrib['checked'] == 'Checked':
                            control.checked = True
    return workbook
